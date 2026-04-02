#!/usr/bin/env python3
"""
Sales Dashboard Generator
Reads the Excel data file and generates net_sales_dashboard.html + index.html
Usage: python3 update_dashboard.py <excel_file> <data_date> <output_dir>
Example: python3 update_dashboard.py "Net Sales.xlsx" "03/30/2026" "./Sales Dashboard - Web"
"""

import sys, os, re
from datetime import datetime, timedelta
from collections import defaultdict, OrderedDict
import openpyxl

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION — update these when targets/static info changes
# ═══════════════════════════════════════════════════════════════════════════════
TARGETS = {
    'gpd_annual': 228, 'gpd_monthly': 19,
    'hld_annual': 372, 'hld_monthly': 31,
    'wrg_annual': 216, 'wrg_monthly': 18,
    'combined_annual': 600, 'combined_monthly': 50,
}

# Static avg prices / $/SF per builder (not in the gross sales data reliably)
GPD_BUILDER_STATIC = {
    ('40', 'David Weekley Homes'):  {'psf': '$159', 'price': '$327,327'},
    ('40', 'Perry Homes'):          {'psf': '$183', 'price': '$332,124'},
    ('45', 'Village Builders'):     {'psf': '$166', 'price': '$311,392'},
    ('45', 'Westin Homes'):         {'psf': '$159', 'price': '$434,378'},
    ('50', 'Perry Homes'):          {'psf': '$176', 'price': '$450,438'},
    ('50', 'David Weekley Homes'):  {'psf': '$158', 'price': '$395,108'},
    ('60', 'Westin Homes'):         {'psf': '$154', 'price': '$571,279'},
    ('60', 'Sitterle Homes'):       {'psf': '$197', 'price': '$594,129'},
    ('60', 'David Weekley Homes'):  {'psf': '$181', 'price': '$512,440'},
    ('60', 'Shea Homes'):           {'psf': '$179', 'price': '$660,040'},
}
GPD_LOT_TOTALS_STATIC = {
    '40': {'psf': '$171', 'price': '$329,689'},
    '45': {'psf': '$163', 'price': '$362,777'},
    '50': {'psf': '$167', 'price': '$421,376'},
    '60': {'psf': '$164', 'price': '$571K'},
}
GPD_GRAND_STATIC = {'psf': '$166', 'price': '~$392K'}
GPD_AVG_PRICE = '~$393K'

HLD_BUILDER_STATIC = {
    ('40', 'Lennar Homes'): {'price': '$250,088'},
    ('45', 'Lennar Homes'): {'price': '$258,490'},
    ('50', 'Lennar Homes'): {'price': '$302,331'},
}
HLD_GRAND_STATIC = {'price': '~$267K'}
HLD_AVG_PRICE = '$260K'

WRG_BUILDER_STATIC = {
    ('40', 'Lennar Homes'):           {'price': '$229,637'},
    ('40', 'CastleRock Communities'): {'price': '$278,942'},
    ('45', 'CastleRock Communities'): {'price': '$324,146'},
    ('45', 'K. Hovnanian'):           {'price': '$320,086'},
    ('45', 'Lennar Homes'):           {'price': '$224,540'},
    ('45', 'Coventry Homes'):         {'price': '$279,993'},
    ('50', 'K. Hovnanian'):           {'price': '$356,611'},
    ('50', 'CastleRock Communities'): {'price': '$338,906'},
    ('50', 'Lennar Homes'):           {'price': '$289,310'},
    ('50', 'Coventry Homes'):         {'price': '$310,368'},
}
WRG_AVG_PRICE = '~$305K'

# GPD target paces per builder
GPD_TARGET_PACE = {
    ('40', 'David Weekley Homes'): 2.0,
    ('40', 'Perry Homes'): 2.0,
    ('45', 'Village Builders'): 4.0,
    ('45', 'Westin Homes'): 2.0,
    ('50', 'Perry Homes'): 3.0,
    ('50', 'David Weekley Homes'): 3.0,
    ('60', 'Westin Homes'): 0.75,
    ('60', 'Sitterle Homes'): 0.75,
    ('60', 'David Weekley Homes'): 0.75,
    ('60', 'Shea Homes'): 0.75,
}

# Builder order for GPD builder-month data
GPD_BM_ORDER = [
    ('40', 'DWH', 'David Weekley Homes'),
    ('40', 'Perry', 'Perry Homes'),
    ('45', 'Village', 'Village Builders'),
    ('45', 'Westin', 'Westin Homes'),
    ('50', 'Perry', 'Perry Homes'),
    ('50', 'DWH', 'David Weekley Homes'),
    ('60', 'Westin', 'Westin Homes'),
    ('60', 'Sitterle', 'Sitterle Homes'),
    ('60', 'DWH', 'David Weekley Homes'),
    ('60', 'Shea', 'Shea Homes'),
]

HLD_BM_ORDER = [
    ('40', 'Lennar', 'Lennar Homes'),
    ('45', 'Lennar', 'Lennar Homes'),
    ('50', 'Lennar', 'Lennar Homes'),
]

WRG_BM_ORDER = [
    ('40', 'Lennar', 'Lennar Homes'),
    ('40', 'CastleRock', 'CastleRock Communities'),
    ('45', 'CastleRock', 'CastleRock Communities'),
    ('45', 'K.Hov', 'K. Hovnanian'),
    ('45', 'Lennar', 'Lennar Homes'),
    ('45', 'Coventry', 'Coventry Homes'),
    ('50', 'K.Hov', 'K. Hovnanian'),
    ('50', 'CastleRock', 'CastleRock Communities'),
    ('50', 'Lennar', 'Lennar Homes'),
    ('50', 'Coventry', 'Coventry Homes'),
]

# GPD pace chart labels and order
GPD_PACE_LABELS = ['DWH 40','Perry 40','Village 45','Westin 45','Perry 50','DWH 50','Westin 60','Sitterle 60']
GPD_PACE_ORDER = [
    ('40','David Weekley Homes'), ('40','Perry Homes'),
    ('45','Village Builders'), ('45','Westin Homes'),
    ('50','Perry Homes'), ('50','David Weekley Homes'),
    ('60','Westin Homes'), ('60','Sitterle Homes'),
]

WRG_PACE_LABELS = ['Lennar 40', 'CastleRock 40', 'CR 45', 'KHov 45', 'Coventry 45', 'KHov 50', 'CR 50', 'Coventry 50']
WRG_PACE_ORDER = [
    ('40','Lennar Homes'), ('40','CastleRock Communities'),
    ('45','CastleRock Communities'), ('45','K. Hovnanian'), ('45','Coventry Homes'),
    ('50','K. Hovnanian'), ('50','CastleRock Communities'), ('50','Coventry Homes'),
]

# TGP first sale month
TGP_START = (2023, 10)  # Oct 2023
# HLD first sale month
HLD_START = (2024, 6)   # Jun 2024
# WRG - we show from Jan 2024

# Badge colors for lot types
BADGE_COLORS = {'40': 'badge-blue', '45': 'badge-green', '50': 'badge-orange', '60': 'badge-red'}

# ═══════════════════════════════════════════════════════════════════════════════
# DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def parse_date(val):
    """Parse a date value from Excel cell - could be datetime or string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ['%m/%d/%y', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None

def get_month_key(dt):
    """Return (year, month) tuple from datetime."""
    return (dt.year, dt.month)

def month_label(ym):
    """Convert (year, month) to 'Mon YY' format."""
    names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    return f"{names[ym[1]-1]} {str(ym[0])[-2:]}"

def get_week_start(dt):
    """Get Monday of the week containing dt."""
    return dt - timedelta(days=dt.weekday())

def week_label(monday_dt):
    """Format week label as 'Mon D'."""
    names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    return f"{names[monday_dt.month-1]} {monday_dt.day}"

def extract_gross_sales(ws):
    """Extract gross sales from a worksheet. Returns list of dicts."""
    sales = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c+1).value for c in range(83)]
        sale_date = parse_date(row[16])
        if sale_date is None:
            continue
        lot_type = row[4]
        if lot_type is None:
            continue
        lot_type = str(int(lot_type)) if isinstance(lot_type, (int, float)) else str(lot_type).strip()
        builder = str(row[8]).strip() if row[8] else 'Unknown'
        sale_price = row[18]
        sqft = row[44]
        sales.append({
            'date': sale_date,
            'lot_type': lot_type,
            'builder': builder,
            'price': sale_price,
            'sqft': sqft,
            'address': str(row[9] or ''),
        })
    return sales

def extract_cancellations(ws):
    """Extract cancellations from a worksheet. Returns list of dicts."""
    cancels = []
    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c+1).value for c in range(18)]
        cancel_date = parse_date(row[10])
        if cancel_date is None:
            continue
        lot_type = row[7]
        if lot_type is None:
            continue
        lot_type = str(int(lot_type)) if isinstance(lot_type, (int, float)) else str(lot_type).strip()
        builder = str(row[6]).strip() if row[6] else 'Unknown'
        sale_date_str = str(row[8] or '')
        cancel_reason = str(row[11] or '')
        address = str(row[1] or '')
        cancels.append({
            'date': cancel_date,
            'lot_type': lot_type,
            'builder': builder,
            'sale_date_str': sale_date_str,
            'reason': cancel_reason,
            'address': address,
        })
    return cancels

def generate_month_range(start_ym, end_ym):
    """Generate list of (year, month) tuples from start to end inclusive."""
    months = []
    y, m = start_ym
    while (y, m) <= end_ym:
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months

def compute_monthly_net(gross_sales, cancellations, month_range):
    """Compute net sales per month. Returns dict {(y,m): net_count}."""
    gross_by_month = defaultdict(int)
    for s in gross_sales:
        ym = get_month_key(s['date'])
        if ym in [(y, m) for y, m in month_range]:
            gross_by_month[ym] += 1

    cancel_by_month = defaultdict(int)
    for c in cancellations:
        ym = get_month_key(c['date'])
        if ym in [(y, m) for y, m in month_range]:
            cancel_by_month[ym] += 1

    result = {}
    for ym in month_range:
        result[ym] = gross_by_month[ym] - cancel_by_month[ym]
    return result

def compute_builder_month_net(gross_sales, cancellations, month_range):
    """Compute net sales per (lot_type, builder) per month."""
    gross = defaultdict(lambda: defaultdict(int))
    for s in gross_sales:
        ym = get_month_key(s['date'])
        key = (s['lot_type'], s['builder'])
        gross[key][ym] += 1

    cancel = defaultdict(lambda: defaultdict(int))
    for c in cancellations:
        ym = get_month_key(c['date'])
        key = (c['lot_type'], c['builder'])
        cancel[key][ym] += 1

    # Get all builder-lot combos
    all_keys = set(gross.keys()) | set(cancel.keys())
    result = {}
    for key in all_keys:
        result[key] = {}
        for ym in month_range:
            result[key][ym] = gross[key].get(ym, 0) - cancel[key].get(ym, 0)
    return result

def compute_weekly_data(gross_sales, cancellations, year=2026):
    """Compute weekly gross/cancel/net for a given year, Monday-start weeks."""
    # Filter to year
    year_gross = [s for s in gross_sales if s['date'].year == year]
    year_cancel = [c for c in cancellations if c['date'].year == year]

    # Exclude Dec 29 week (belongs to prior year visually)
    dec29_cutoff = datetime(year, 1, 1)
    # Find first Monday of Jan that's >= Jan 1
    first_monday = datetime(year, 1, 1)
    while first_monday.weekday() != 0:
        first_monday += timedelta(days=1)
    # If Jan 1 is not Monday, the first full week starts on first Monday
    # But sales from Jan 1 to first Sunday belong to the Dec 29 week - exclude them
    # Actually, get_week_start will place them in the right week

    gross_by_week = defaultdict(int)
    cancel_by_week = defaultdict(int)
    gross_detail = defaultdict(lambda: defaultdict(int))  # {(lot,builder): {week_monday: count}}
    cancel_detail = defaultdict(lambda: defaultdict(int))

    for s in year_gross:
        ws = get_week_start(s['date'])
        if ws.year < year:  # Dec 29 week - skip
            continue
        gross_by_week[ws] += 1
        gross_detail[(s['lot_type'], s['builder'])][ws] += 1

    for c in year_cancel:
        ws = get_week_start(c['date'])
        if ws.year < year:
            continue
        cancel_by_week[ws] += 1
        cancel_detail[(c['lot_type'], c['builder'])][ws] += 1

    # Get ALL weeks from first Monday of year to last week with data
    active_weeks = set(gross_by_week.keys()) | set(cancel_by_week.keys())
    if not active_weeks:
        return {'weeks': [], 'labels': [], 'gross': [], 'cancel': [], 'builder_weekly': {}}

    first_monday = datetime(year, 1, 1)
    while first_monday.weekday() != 0:
        first_monday += timedelta(days=1)
    last_week = max(active_weeks)

    all_weeks = []
    current_week = first_monday
    while current_week <= last_week:
        all_weeks.append(current_week)
        current_week += timedelta(days=7)

    week_labels = [week_label(w) for w in all_weeks]
    week_gross = [gross_by_week[w] for w in all_weeks]
    week_cancel = [-cancel_by_week[w] for w in all_weeks]  # negative

    # Builder detail
    all_builder_keys = set(gross_detail.keys()) | set(cancel_detail.keys())
    builder_weekly = {}
    for bk in all_builder_keys:
        g = [gross_detail[bk].get(w, 0) for w in all_weeks]
        c = [-cancel_detail[bk].get(w, 0) for w in all_weeks]
        n = [g[i] + c[i] for i in range(len(all_weeks))]
        builder_weekly[bk] = {'g': g, 'c': c, 'n': n}

    return {
        'weeks': all_weeks,
        'labels': week_labels,
        'gross': week_gross,
        'cancel': week_cancel,
        'builder_weekly': builder_weekly,
    }

def compute_all_time_totals(gross_sales, cancellations):
    """Compute all-time gross/cancel/net per (lot_type, builder)."""
    gross = defaultdict(int)
    cancel = defaultdict(int)
    for s in gross_sales:
        gross[(s['lot_type'], s['builder'])] += 1
    for c in cancellations:
        cancel[(c['lot_type'], c['builder'])] += 1
    all_keys = set(gross.keys()) | set(cancel.keys())
    result = {}
    for key in all_keys:
        g = gross[key]
        c = cancel[key]
        result[key] = {'tg': g, 'tc': -c, 'tn': g - c}
    return result

def compute_avg_pace(net_by_month, months_list):
    """Compute average pace (net/month) for a range of months."""
    if not months_list:
        return 0.0
    total = sum(net_by_month.get(m, 0) for m in months_list)
    return round(total / len(months_list), 2)


# ═══════════════════════════════════════════════════════════════════════════════
# HTML TEMPLATE GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def js_array(arr):
    """Convert Python list to JavaScript array string."""
    parts = []
    for v in arr:
        if v is None:
            parts.append('null')
        elif isinstance(v, float):
            parts.append(f'{v:.2f}' if v != int(v) else str(int(v)))
        else:
            parts.append(str(v))
    return '[' + ','.join(parts) + ']'

def js_str_array(arr):
    """Convert list of strings to JS array."""
    return '[' + ','.join(f"'{s}'" for s in arr) + ']'

def fmt_num(n):
    """Format number with commas."""
    return f'{n:,}'

def fmt_pct(n, total):
    """Format percentage."""
    if total == 0:
        return '0.0%'
    return f'{(n/total)*100:.1f}%'

def fmt_pace(val):
    """Format pace value."""
    if val == 0:
        return '0.00'
    return f'{val:.2f}'


def generate_html(data, data_date):
    """Generate the complete HTML dashboard."""
    d = data  # shorthand

    # Determine current year and months available
    current_year = int(data_date.split('/')[-1])
    current_month = int(data_date.split('/')[0])
    ytd_months_count = current_month  # Jan through current month
    month_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    ytd_label = f"Jan–{month_names[current_month-1]}"

    # ── BUILD KPIs ──────────────────────────────────────────────────────────

    gpd_lop = d['gpd_total_net']
    hld_lop = d['hld_total_net']
    wrg_lop = d['wrg_total_net']
    combined_lop = gpd_lop + hld_lop

    gpd_2026_net = d['gpd_ytd_net']
    hld_2026_net = d['hld_ytd_net']
    wrg_2026_net = d['wrg_ytd_net']

    gpd_avg_mo = round(gpd_2026_net / ytd_months_count, 1)
    hld_avg_mo = round(hld_2026_net / ytd_months_count, 1)
    wrg_avg_mo = round(wrg_2026_net / ytd_months_count, 1)
    gp_avg_mo = round((gpd_2026_net + hld_2026_net) / ytd_months_count, 1)

    # 2026 cancellations
    canc_2026 = d['gpd_canc_2026'] + d['hld_canc_2026'] + d['wrg_canc_2026']

    # ── GPD Builder Summary rows ────────────────────────────────────────────
    def build_gpd_summary():
        rows_html = ''
        lot_groups = OrderedDict()
        for lot, short, full in GPD_BM_ORDER:
            if lot not in lot_groups:
                lot_groups[lot] = []
            bm_net = d['gpd_bm_net'].get((lot, full), {})
            # Compute paces
            months_2024 = [m for m in d['gpd_months'] if m[0] == 2024]
            months_2025 = [m for m in d['gpd_months'] if m[0] == 2025]
            months_2026 = [m for m in d['gpd_months'] if m[0] == current_year and m[1] <= current_month]
            pace_24 = compute_avg_pace(bm_net, months_2024)
            pace_25 = compute_avg_pace(bm_net, months_2025)
            pace_26 = compute_avg_pace(bm_net, months_2026)
            total_sales = d['gpd_at'].get((lot, full), {}).get('tn', 0)
            target = GPD_TARGET_PACE.get((lot, full), 0)
            static = GPD_BUILDER_STATIC.get((lot, full), {})
            lot_groups[lot].append({
                'builder': full, 'pace24': pace_24, 'pace25': pace_25,
                'pace26': pace_26, 'target': target, 'total': total_sales,
                'psf': static.get('psf', '—'), 'price': static.get('price', '—'),
            })

        grand_total = gpd_lop
        grand_pace24 = sum(r['pace24'] for lt in lot_groups for r in lot_groups[lt])
        grand_pace25 = sum(r['pace25'] for lt in lot_groups for r in lot_groups[lt])
        grand_pace26 = sum(r['pace26'] for lt in lot_groups for r in lot_groups[lt])
        grand_target = sum(r['target'] for lt in lot_groups for r in lot_groups[lt])

        for lot, builders in lot_groups.items():
            badge = BADGE_COLORS.get(lot, 'badge-blue')
            for b in builders:
                pct = fmt_pct(b['total'], grand_total)
                rows_html += f'          <tr><td><span class="badge {badge}">{lot}</span></td><td>{b["builder"]}</td>'
                rows_html += f'<td class="right">{fmt_pace(b["pace24"])}</td><td class="right">{fmt_pace(b["pace25"])}</td>'
                rows_html += f'<td class="right">{fmt_pace(b["pace26"])}</td><td class="right">{b["target"]}</td>'
                rows_html += f'<td class="right">{b["psf"]}</td><td class="right">{b["price"]}</td>'
                rows_html += f'<td class="right">{b["total"]}</td><td class="right">{pct}</td></tr>\n'
            # Subtotal
            lt_total = sum(b['total'] for b in builders)
            lt_pace24 = sum(b['pace24'] for b in builders)
            lt_pace25 = sum(b['pace25'] for b in builders)
            lt_pace26 = sum(b['pace26'] for b in builders)
            lt_target = sum(b['target'] for b in builders)
            lt_static = GPD_LOT_TOTALS_STATIC.get(lot, {})
            rows_html += f'          <tr class="total-row"><td colspan="2">{lot} Total</td>'
            rows_html += f'<td class="right">{fmt_pace(lt_pace24)}</td><td class="right">{fmt_pace(lt_pace25)}</td>'
            rows_html += f'<td class="right">{fmt_pace(lt_pace26)}</td><td class="right">{lt_target}</td>'
            rows_html += f'<td class="right"><strong>{lt_static.get("psf","—")}</strong></td><td class="right">{lt_static.get("price","—")}</td>'
            rows_html += f'<td class="right">{lt_total}</td><td class="right"><strong>{fmt_pct(lt_total, grand_total)}</strong></td></tr>\n'

        rows_html += f'          <tr class="total-row" style="background:#d6e4ff"><td colspan="2"><strong>GRAND TOTAL</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_pace(grand_pace24)}</strong></td><td class="right"><strong>{fmt_pace(grand_pace25)}</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_pace(grand_pace26)}</strong></td><td class="right"><strong>{grand_target}</strong></td>'
        rows_html += f'<td class="right"><strong>{GPD_GRAND_STATIC["psf"]}</strong></td><td class="right"><strong>{GPD_GRAND_STATIC["price"]}</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_num(grand_total)}</strong></td><td class="right"><strong>100%</strong></td></tr>\n'
        return rows_html

    def build_hld_summary():
        rows_html = ''
        months_2024 = [m for m in d['hld_months'] if m[0] == 2024]
        months_2025 = [m for m in d['hld_months'] if m[0] == 2025]
        months_2026 = [m for m in d['hld_months'] if m[0] == current_year and m[1] <= current_month]
        # T12 = trailing 12 months
        all_months_sorted = sorted(d['hld_months'])
        t12_months = all_months_sorted[-12:] if len(all_months_sorted) >= 12 else all_months_sorted

        grand_total = hld_lop
        entries = []
        for lot, short, full in HLD_BM_ORDER:
            bm_net = d['hld_bm_net'].get((lot, full), {})
            pace24 = compute_avg_pace(bm_net, months_2024)
            pace25 = compute_avg_pace(bm_net, months_2025)
            pace26 = compute_avg_pace(bm_net, months_2026)
            t12 = compute_avg_pace(bm_net, t12_months)
            total_sales = d['hld_at'].get((lot, full), {}).get('tn', 0)
            static = HLD_BUILDER_STATIC.get((lot, full), {})
            badge = BADGE_COLORS.get(lot, 'badge-blue')
            pct = fmt_pct(total_sales, grand_total)
            rows_html += f'          <tr><td><span class="badge {badge}">{lot}</span></td><td>{full}</td>'
            rows_html += f'<td class="right">{fmt_pace(pace24)}</td><td class="right">{fmt_pace(pace25)}</td>'
            rows_html += f'<td class="right">{fmt_pace(pace26)}</td><td class="right">{fmt_pace(t12)}</td>'
            rows_html += f'<td class="right">{static.get("price","—")}</td>'
            rows_html += f'<td class="right">{total_sales}</td><td class="right">{pct}</td></tr>\n'
            entries.append({'pace24': pace24, 'pace25': pace25, 'pace26': pace26, 't12': t12, 'total': total_sales})

        gp24 = sum(e['pace24'] for e in entries)
        gp25 = sum(e['pace25'] for e in entries)
        gp26 = sum(e['pace26'] for e in entries)
        gt12 = sum(e['t12'] for e in entries)
        rows_html += f'          <tr class="total-row" style="background:#e6f4ea"><td colspan="2"><strong>TOTAL</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_pace(gp24)}</strong></td><td class="right"><strong>{fmt_pace(gp25)}</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_pace(gp26)}</strong></td><td class="right"><strong>{fmt_pace(gt12)}</strong></td>'
        rows_html += f'<td class="right"><strong>{HLD_GRAND_STATIC["price"]}</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_num(grand_total)}</strong></td><td class="right"><strong>100%</strong></td></tr>\n'
        return rows_html

    def build_wrg_summary():
        rows_html = ''
        months_2024 = generate_month_range((2024,1),(2024,12))
        months_2025 = generate_month_range((2025,1),(2025,12))
        months_2026 = generate_month_range((current_year,1),(current_year,current_month))
        all_months_sorted = sorted(d['wrg_months'])
        t12_months = all_months_sorted[-12:] if len(all_months_sorted) >= 12 else all_months_sorted

        grand_total = wrg_lop
        lot_groups = OrderedDict()
        for lot, short, full in WRG_BM_ORDER:
            if lot not in lot_groups:
                lot_groups[lot] = []
            bm_net = d['wrg_bm_net'].get((lot, full), {})
            pace24 = compute_avg_pace(bm_net, months_2024)
            pace25 = compute_avg_pace(bm_net, months_2025)
            pace26 = compute_avg_pace(bm_net, months_2026)
            t12 = compute_avg_pace(bm_net, t12_months)
            total_sales = d['wrg_at'].get((lot, full), {}).get('tn', 0)
            static = WRG_BUILDER_STATIC.get((lot, full), {})
            lot_groups[lot].append({
                'builder': full, 'pace24': pace24, 'pace25': pace25,
                'pace26': pace26, 't12': t12, 'total': total_sales,
                'price': static.get('price', '—'),
            })

        for lot, builders in lot_groups.items():
            badge = BADGE_COLORS.get(lot, 'badge-blue')
            for b in builders:
                pct = fmt_pct(b['total'], grand_total)
                rows_html += f'          <tr><td><span class="badge {badge}">{lot}</span></td><td>{b["builder"]}</td>'
                rows_html += f'<td class="right">{fmt_pace(b["pace24"])}</td><td class="right">{fmt_pace(b["pace25"])}</td>'
                rows_html += f'<td class="right">{fmt_pace(b["pace26"])}</td><td class="right">{fmt_pace(b["t12"])}</td>'
                rows_html += f'<td class="right">{b["price"]}</td>'
                rows_html += f'<td class="right">{b["total"]}</td><td class="right">{pct}</td></tr>\n'
            lt_total = sum(b['total'] for b in builders)
            lt_pace24 = sum(b['pace24'] for b in builders)
            lt_pace25 = sum(b['pace25'] for b in builders)
            lt_pace26 = sum(b['pace26'] for b in builders)
            lt_t12 = sum(b['t12'] for b in builders)
            rows_html += f'          <tr class="total-row"><td colspan="2">{lot} Total</td>'
            rows_html += f'<td class="right">{fmt_pace(lt_pace24)}</td><td class="right">{fmt_pace(lt_pace25)}</td>'
            rows_html += f'<td class="right">{fmt_pace(lt_pace26)}</td><td class="right">{fmt_pace(lt_t12)}</td>'
            rows_html += f'<td class="right">—</td>'  # no combined price for WRG lot totals
            rows_html += f'<td class="right">{lt_total}</td><td class="right"><strong>{fmt_pct(lt_total, grand_total)}</strong></td></tr>\n'

        gp24 = sum(b['pace24'] for lt in lot_groups for b in lot_groups[lt])
        gp25 = sum(b['pace25'] for lt in lot_groups for b in lot_groups[lt])
        gp26 = sum(b['pace26'] for lt in lot_groups for b in lot_groups[lt])
        gt12 = sum(b['t12'] for lt in lot_groups for b in lot_groups[lt])
        rows_html += f'          <tr class="total-row" style="background:#ffe4e4"><td colspan="2"><strong>GRAND TOTAL</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_pace(gp24)}</strong></td><td class="right"><strong>{fmt_pace(gp25)}</strong></td>'
        rows_html += f'<td class="right"><strong>{fmt_pace(gp26)}</strong></td><td class="right"><strong>{fmt_pace(gt12)}</strong></td>'
        rows_html += f'<td class="right">—</td>'
        rows_html += f'<td class="right"><strong>{fmt_num(grand_total)}</strong></td><td class="right"><strong>100%</strong></td></tr>\n'
        return rows_html

    # ── Recent cancellations ────────────────────────────────────────────────
    def build_recent_cancellations():
        recent = sorted(d['gpd_recent_cancels'], key=lambda x: x['date'], reverse=True)[:5]
        rows_html = ''
        for c in recent:
            badge = BADGE_COLORS.get(c['lot_type'], 'badge-blue')
            sd = c['sale_date_str']
            cd = c['date'].strftime('%m/%d/%y')
            rows_html += f'          <tr><td>{c["address"]}</td><td>{c["builder"]}</td>'
            rows_html += f'<td><span class="badge {badge}">{c["lot_type"]}</span></td>'
            rows_html += f'<td>{sd}</td><td>{cd}</td><td>{c["reason"]}</td></tr>\n'
        return rows_html

    # ── Build JS data arrays ────────────────────────────────────────────────

    # Shared months (Jan 2024 – current)
    shared_range = generate_month_range((2024,1), (current_year, current_month))
    shared_labels = [month_label(m) for m in shared_range]

    tgp_shared = [d['gpd_net_monthly'].get(m, 0) for m in shared_range]
    hld_shared = [d['hld_net_monthly'].get(m, 0) for m in shared_range]
    wrg_shared = [d['wrg_net_monthly'].get(m, 0) for m in shared_range]

    # TGP full monthly
    tgp_month_labels = [month_label(m) for m in d['gpd_months']]
    tgp_month_data = [d['gpd_net_monthly'].get(m, 0) for m in d['gpd_months']]

    # HLD full monthly
    hld_month_labels = [month_label(m) for m in d['hld_months']]
    hld_month_data = [d['hld_net_monthly'].get(m, 0) for m in d['hld_months']]

    # WRG monthly - skip months with 0 sales
    wrg_all_months = generate_month_range((2024,1), (current_year, current_month))
    wrg_month_labels = []
    wrg_month_data_arr = []
    for m in wrg_all_months:
        val = d['wrg_net_monthly'].get(m, 0)
        if val == 0 and m[0] < current_year:
            continue  # skip zero months (like Jul 2025)
        wrg_month_labels.append(month_label(m))
        wrg_month_data_arr.append(val)

    # Pace arrays (same as shared)
    pace_months = shared_range

    # Builder-month data
    def build_bm_js(bm_order, bm_net, months):
        lines = []
        for lot, short, full in bm_order:
            key = (lot, full)
            vals = [bm_net.get(key, {}).get(m, 0) for m in months]
            lines.append(f"  '{lot}ft – {short}': {js_array(vals)}")
        return '{\n' + ',\n'.join(lines) + '\n}'

    tgp_bm_js = build_bm_js(GPD_BM_ORDER, d['gpd_bm_net'], d['gpd_months'])
    hld_bm_js = build_bm_js(HLD_BM_ORDER, d['hld_bm_net'], d['hld_months'])
    wrg_bm_months = generate_month_range((2024,1), (current_year, current_month))
    wrg_bm_js = build_bm_js(WRG_BM_ORDER, d['wrg_bm_net'], wrg_bm_months)

    # YOY data
    def build_yoy(net_monthly, years, start_month=1):
        result = {}
        for yr in years:
            vals = []
            for mo in range(1, 13):
                ym = (yr, mo)
                if ym in net_monthly and (yr < current_year or mo <= current_month):
                    vals.append(net_monthly[ym])
                else:
                    vals.append(None)
            result[yr] = vals
        return result

    # GPD YOY - includes 2023 (only Oct-Dec)
    gpd_yoy = {}
    gpd_yoy[2023] = [None]*9 + [d['gpd_net_monthly'].get((2023,10),0), d['gpd_net_monthly'].get((2023,11),0), d['gpd_net_monthly'].get((2023,12),0)]
    gpd_yoy[2024] = [d['gpd_net_monthly'].get((2024,m),0) for m in range(1,13)]
    gpd_yoy[2025] = [d['gpd_net_monthly'].get((2025,m),0) for m in range(1,13)]
    gpd_yoy[2026] = [d['gpd_net_monthly'].get((current_year,m),0) if m <= current_month else None for m in range(1,13)]

    hld_yoy = {}
    hld_yoy[2024] = [None]*5 + [d['hld_net_monthly'].get((2024,m),0) for m in range(6,13)]
    hld_yoy[2025] = [d['hld_net_monthly'].get((2025,m),0) for m in range(1,13)]
    hld_yoy[2026] = [d['hld_net_monthly'].get((current_year,m),0) if m <= current_month else None for m in range(1,13)]

    wrg_yoy = {}
    wrg_yoy[2024] = [d['wrg_net_monthly'].get((2024,m),0) for m in range(1,13)]
    wrg_yoy[2025] = [d['wrg_net_monthly'].get((2025,m),0) for m in range(1,13)]
    wrg_yoy[2026] = [d['wrg_net_monthly'].get((current_year,m),0) if m <= current_month else None for m in range(1,13)]

    # Combined GP YOY
    combined_yoy = {}
    for yr in [2023, 2024, 2025, 2026]:
        gvals = gpd_yoy.get(yr, [None]*12)
        hvals = hld_yoy.get(yr, [None]*12)
        combined_yoy[yr] = []
        for i in range(12):
            g, h = gvals[i], hvals[i] if yr in hld_yoy else None
            if g is None and h is None:
                combined_yoy[yr].append(None)
            else:
                combined_yoy[yr].append((g or 0) + (h or 0))

    def yoy_js(data):
        lines = []
        for yr, vals in sorted(data.items()):
            lines.append(f"    {yr}: {js_array(vals)}")
        return '{\n' + ',\n'.join(lines) + ',\n  }'

    # Doughnut chart data (LOP totals by lot type)
    # GPD lot totals
    gpd_lot_totals = defaultdict(int)
    for (lot, builder), info in d['gpd_at'].items():
        gpd_lot_totals[lot] += info['tn']

    hld_lot_totals = defaultdict(int)
    for (lot, builder), info in d['hld_at'].items():
        hld_lot_totals[lot] += info['tn']

    wrg_lot_totals = defaultdict(int)
    for (lot, builder), info in d['wrg_at'].items():
        wrg_lot_totals[lot] += info['tn']

    # Weekly data
    gpd_wk = d['gpd_weekly']
    hld_wk = d['hld_weekly']
    wrg_wk = d['wrg_weekly']

    # Weekly detail rows JS
    def build_weekly_rows_js(weekly_data, at_totals, bm_order):
        rows = []
        for lot, short, full in bm_order:
            key = (lot, full)
            wk = weekly_data['builder_weekly'].get(key, None)
            at = at_totals.get(key, {'tg':0,'tc':0,'tn':0})
            n_weeks = len(weekly_data['labels'])
            if wk:
                g = wk['g']
                c = wk['c']
                n = wk['n']
            else:
                g = [0] * n_weeks
                c = [0] * n_weeks
                n = [0] * n_weeks
            rows.append(f'  {{lot:{lot},bld:"{full}",g:{js_array(g)},c:{js_array(c)},n:{js_array(n)},tg:{at["tg"]},tc:{at["tc"]},tn:{at["tn"]}}}')
        return '[\n' + ',\n'.join(rows) + '\n]'

    gpd_wk_rows_js = build_weekly_rows_js(gpd_wk, d['gpd_at'], GPD_BM_ORDER)
    hld_wk_rows_js = build_weekly_rows_js(hld_wk, d['hld_at'], HLD_BM_ORDER)
    wrg_wk_rows_js = build_weekly_rows_js(wrg_wk, d['wrg_at'], WRG_BM_ORDER)

    # GPD pace chart data per builder
    months_2024 = generate_month_range((2024,1),(2024,12))
    months_2025 = generate_month_range((2025,1),(2025,12))
    months_2026 = generate_month_range((current_year,1),(current_year,current_month))

    gpd_pace24 = [compute_avg_pace(d['gpd_bm_net'].get(k,{}), months_2024) for k in GPD_PACE_ORDER]
    gpd_pace25 = [compute_avg_pace(d['gpd_bm_net'].get(k,{}), months_2025) for k in GPD_PACE_ORDER]
    gpd_pace26 = [compute_avg_pace(d['gpd_bm_net'].get(k,{}), months_2026) for k in GPD_PACE_ORDER]

    # HLD pace chart
    hld_months_2024 = [m for m in d['hld_months'] if m[0]==2024]
    hld_pace24 = [compute_avg_pace(d['hld_bm_net'].get((lot,full),{}), hld_months_2024) for lot,_,full in HLD_BM_ORDER]
    hld_pace25 = [compute_avg_pace(d['hld_bm_net'].get((lot,full),{}), months_2025) for lot,_,full in HLD_BM_ORDER]
    hld_pace26 = [compute_avg_pace(d['hld_bm_net'].get((lot,full),{}), months_2026) for lot,_,full in HLD_BM_ORDER]

    # WRG pace chart
    wrg_pace24 = [compute_avg_pace(d['wrg_bm_net'].get(k,{}), months_2024) for k in WRG_PACE_ORDER]
    wrg_pace25 = [compute_avg_pace(d['wrg_bm_net'].get(k,{}), months_2025) for k in WRG_PACE_ORDER]
    wrg_pace26 = [compute_avg_pace(d['wrg_bm_net'].get(k,{}), months_2026) for k in WRG_PACE_ORDER]

    # Overview pace chart
    gpd_full_pace24 = round(sum(d['gpd_net_monthly'].get(m,0) for m in months_2024)/12, 1)
    gpd_full_pace25 = round(sum(d['gpd_net_monthly'].get(m,0) for m in months_2025)/12, 1)
    hld_full_pace24 = round(sum(d['hld_net_monthly'].get(m,0) for m in hld_months_2024)/len(hld_months_2024), 1) if hld_months_2024 else 0
    hld_full_pace25 = round(sum(d['hld_net_monthly'].get(m,0) for m in months_2025)/12, 1)

    # Cancellation totals
    gpd_canc_lop = d['gpd_canc_total']
    hld_canc_lop = d['hld_canc_total']
    wrg_canc_lop = d['wrg_canc_total']
    total_canc = gpd_canc_lop + hld_canc_lop + wrg_canc_lop

    # Gross sales totals
    gpd_gross_lop = d['gpd_gross_total']
    hld_gross_lop = d['hld_gross_total']
    wrg_gross_lop = d['wrg_gross_total']

    # Mix chart (GPD vs HLD at current LOP — uses older static numbers for doughnut)
    # Actually compute from data

    # ── STARTS DATA ─────────────────────────────────────────────────────────
    # Starts data is NOT in the Gross Sales/Cancellations sheets — it's in separate sheets
    # We'll pass through from the dashboard as-is since those sheets have different structure
    # For now, keep starts as static JS (read from the existing dashboard)

    # ═════════════════════════════════════════════════════════════════════════
    # HTML OUTPUT
    # ═════════════════════════════════════════════════════════════════════════

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,400;0,500;0,600;0,700;0,800;1,400;1,600&display=swap" rel="stylesheet">
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>Sales Dashboard — The Grand Prairie · Windrose Green</title>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
  <script src="https://cdnjs.cloudflare.com/ajax/libs/chartjs-plugin-datalabels/2.2.0/chartjs-plugin-datalabels.min.js"></script>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}

    body {{
      font-family: 'Plus Jakarta Sans', 'Segoe UI', Arial, sans-serif;
      background: #f0f2f5;
      color: #003552;
      min-height: 100vh;
    }}

    /* ── HEADER */
    header {{
      background: linear-gradient(135deg, #003552 0%, #003552 100%);
      color: #fff;
      padding: 28px 40px 22px;
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      box-shadow: 0 4px 20px rgba(0,0,0,.25);
    }}
    header h1 {{ font-size: 1.75rem; font-weight: 700; letter-spacing: -.3px; }}
    header p  {{ font-size: .85rem; color: #a0aec0; margin-top: 4px; }}
    .pulled {{ font-size: .78rem; color: #7c8fb0; }}

    /* ── NAV TABS */
    nav {{
      background: #fff;
      border-bottom: 2px solid #e2e8f0;
      display: flex;
      gap: 0;
      padding: 0 32px;
      overflow-x: auto;
    }}
    nav button {{
      padding: 14px 22px;
      border: none;
      background: none;
      cursor: pointer;
      font-size: .92rem;
      font-weight: 500;
      color: #718096;
      white-space: nowrap;
      border-bottom: 3px solid transparent;
      transition: all .15s;
    }}
    nav button:hover {{ color: #2d3748; }}
    nav button.active {{ color: #003552; border-bottom-color: #003552; font-weight: 700; }}

    /* ── MAIN CONTENT */
    main {{ padding: 28px 32px; max-width: 1400px; margin: 0 auto; }}

    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}

    /* ── KPI CARDS */
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 16px;
      margin-bottom: 28px;
    }}
    .kpi {{
      background: #fff;
      border-radius: 12px;
      padding: 20px 22px;
      box-shadow: 0 1px 6px rgba(0,0,0,.08);
      border-left: 4px solid #003552;
    }}
    .kpi.tgp   {{ border-left-color: #003552; }}
    .kpi.hld   {{ border-left-color: #38a169; }}
    .kpi.wrg   {{ border-left-color: #e53e3e; }}
    .kpi.total {{ border-left-color: #d69e2e; }}
    .kpi label {{ font-size: .72rem; font-weight: 600; text-transform: uppercase; letter-spacing: .6px; color: #718096; }}
    .kpi .value {{ font-size: 2rem; font-weight: 800; margin: 4px 0 2px; color: #111111; }}
    .kpi .sub   {{ font-size: .78rem; color: #a0aec0; }}

    /* ── SECTION TITLES */
    .section-title {{
      font-size: 1.05rem;
      font-weight: 700;
      color: #000 !important;
      margin: 28px 0 14px;
      padding-bottom: 6px;
      border-bottom: 2px solid #e2e8f0;
    }}

    /* ── CHART GRID */
    .chart-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
      margin-bottom: 28px;
    }}
    .chart-grid.single {{ grid-template-columns: 1fr; }}
    @media (max-width: 900px) {{ .chart-grid {{ grid-template-columns: 1fr; }} }}

    .chart-card {{
      background: #fff;
      border-radius: 12px;
      padding: 22px 24px;
      box-shadow: 0 1px 6px rgba(0,0,0,.08);
    }}
    .chart-card h3 {{ font-size: .88rem; font-weight: 700; color: #000 !important; margin-bottom: 16px; text-transform: uppercase; letter-spacing: .5px; }}
    .chart-card canvas {{ max-height: 280px; }}

    /* ── DATA TABLE */
    .table-wrapper {{
      background: #fff;
      border-radius: 12px;
      box-shadow: 0 1px 6px rgba(0,0,0,.08);
      overflow-x: auto;
      margin-bottom: 28px;
    }}
    table {{ width: 100%; border-collapse: separate; border-spacing: 0; font-size: .84rem; }}
    th {{
      background: #003552;
      color: #fff;
      padding: 11px 14px;
      text-align: left;
      font-weight: 600;
      font-size: .78rem;
      letter-spacing: .4px;
      text-transform: uppercase;
      position: sticky;
      top: 0;
      z-index: 2;
    }}
    th:first-child {{
      left: 0;
      z-index: 3;
    }}
    th.right, td.right {{ text-align: right; }}
    td {{ padding: 10px 14px; border-bottom: 1px solid #f0f2f5; }}
    tr:last-child td {{ border-bottom: none; }}
    tr:hover td {{ background: #f7faff; }}
    .total-row td {{ background: #eef2ff; font-weight: 700; }}
    .freeze-col td:first-child,
    .freeze-col th:first-child {{
      position: sticky;
      left: 0;
      z-index: 1;
      background: #003552;
      color: #fff;
      box-shadow: 2px 0 5px rgba(0,0,0,.10);
    }}
    .freeze-col td:first-child {{
      background: #f8faff;
      color: #003552;
      font-weight: 500;
      border-right: 1px solid #e2e8f0;
    }}
    .freeze-col tr:hover td:first-child {{ background: #eef4ff; }}
    .freeze-col .total-row td:first-child {{ background: #dce8ff; }}
    .freeze-col td:nth-child(2),
    .freeze-col th:nth-child(2) {{
      position: sticky;
      left: 64px;
      z-index: 1;
      background: #003552;
      color: #fff;
      box-shadow: 2px 0 5px rgba(0,0,0,.10);
    }}
    .freeze-col td:nth-child(2) {{
      background: #f8faff;
      color: #003552;
      border-right: 2px solid #e2e8f0;
    }}
    .freeze-col tr:hover td:nth-child(2) {{ background: #eef4ff; }}
    .freeze-col .total-row td:nth-child(2) {{ background: #dce8ff; }}
    .badge {{
      display: inline-block;
      padding: 2px 9px;
      border-radius: 999px;
      font-size: .72rem;
      font-weight: 600;
    }}
    .badge-blue   {{ background: #ebf4ff; color: #003552; }}
    .badge-green  {{ background: #f0fff4; color: #276749; }}
    .badge-red    {{ background: #fff5f5; color: #c53030; }}
    .badge-orange {{ background: #fffaf0; color: #c05621; }}

    /* ── PROJECTIONS TABLE */
    .proj-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 20px;
      margin-bottom: 28px;
    }}
    @media (max-width: 800px) {{ .proj-grid {{ grid-template-columns: 1fr; }} }}

    /* ── FOOTER */
    footer {{
      text-align: center;
      padding: 20px;
      font-size: .75rem;
      color: #a0aec0;
    }}
  </style>
</head>
<body>

<header>
  <div style="display:flex;align-items:center;gap:20px">
    <div style="display:flex;align-items:center;gap:10px;flex-shrink:0">
      <svg viewBox="0 0 45 30" width="42" height="30" xmlns="http://www.w3.org/2000/svg">
        <rect x="0"  y="22" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="0"  y="15" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="0"  y="8"  width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="8"  y="22" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="8"  y="15" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="8"  y="8"  width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="8"  y="1"  width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="16" y="22" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="16" y="15" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="16" y="8"  width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="24" y="22" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="24" y="15" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="32" y="22" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="32" y="15" width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="32" y="8"  width="5" height="5" rx="1" fill="#E55500"/>
        <rect x="32" y="1"  width="5" height="5" rx="1" fill="#E55500"/>
      </svg>
      <span style="font-size:1.45rem;font-weight:800;letter-spacing:3px;color:#fff;font-family:'Plus Jakarta Sans',sans-serif">EMBER</span>
    </div>
    <div style="width:1px;height:38px;background:rgba(255,255,255,.25);flex-shrink:0"></div>
    <div>
      <h1>Sales Dashboard</h1>
      <p>The Grand Prairie &nbsp;·&nbsp; Windrose Green</p>
    </div>
  </div>
  <span class="pulled">Data Pulled: {data_date}</span>
</header>

<nav>
  <button class="active" onclick="showTab('overview')">The Grand Prairie</button>
  <button onclick="showTab('tgp')">GPD</button>
  <button onclick="showTab('highlands')">Highlands</button>
  <button onclick="showTab('wrg')">Windrose Green</button>
  <button onclick="showTab('projections')">2026 Projections</button>
  <button onclick="showTab('cancellations')">Cancellations</button>
  <button onclick="showTab('weekly')">Sales by Week</button>
  <button onclick="showTab('starts')">Starts</button>
</nav>

<main>

  <!-- THE GRAND PRAIRIE TAB -->
  <div id="tab-overview" class="tab-panel active">

    <div class="kpi-grid">
      <div class="kpi tgp">
        <label>Total Sales (LOP)</label>
        <div class="value">{fmt_num(combined_lop)}</div>
        <div class="sub">GPD + Highlands combined</div>
      </div>
      <div class="kpi tgp">
        <label>GPD Sales</label>
        <div class="value">{fmt_num(gpd_lop)}</div>
        <div class="sub">Life of Project</div>
      </div>
      <div class="kpi hld">
        <label>Highlands Sales</label>
        <div class="value">{fmt_num(hld_lop)}</div>
        <div class="sub">Life of Project</div>
      </div>
      <div class="kpi total">
        <label>2026 Target</label>
        <div class="value">{TARGETS["combined_annual"]}/yr</div>
        <div class="sub">GPD {TARGETS["gpd_annual"]} · Hld {TARGETS["hld_annual"]}</div>
      </div>
      <div class="kpi tgp">
        <label>GP 2026 Avg/mo</label>
        <div class="value">{gp_avg_mo}</div>
        <div class="sub">{ytd_label} avg · target {TARGETS["combined_monthly"]}</div>
      </div>
      <div class="kpi tgp">
        <label>GPD 2026 Avg/mo</label>
        <div class="value">{gpd_avg_mo}</div>
        <div class="sub">{ytd_label} avg · target {TARGETS["gpd_monthly"]}</div>
      </div>
      <div class="kpi hld">
        <label>Highlands 2026 Avg/mo</label>
        <div class="value">{hld_avg_mo}</div>
        <div class="sub">{ytd_label} avg · target {TARGETS["hld_monthly"]}</div>
      </div>
      <div class="kpi total">
        <label>2026 Cancellations</label>
        <div class="value">{canc_2026}</div>
        <div class="sub">GPD LOP</div>
      </div>
    </div>

    <div class="section-title">Monthly Net Sales — The Grand Prairie ({month_label(shared_range[0])} – {month_label(shared_range[-1])})</div>
    <div class="chart-grid single">
      <div class="chart-card">
        <h3>GPD vs Highlands vs Combined — Net Sales by Month</h3>
        <canvas id="overviewChart"></canvas>
      </div>
    </div>

    <div class="chart-grid">
      <div class="chart-card">
        <h3>Sales Mix by Sub-Community (LOP)</h3>
        <canvas id="mixChart"></canvas>
      </div>
      <div class="chart-card">
        <h3>Avg. Monthly Pace — GPD vs Highlands (2024 · 2025 · 2026 YTD)</h3>
        <canvas id="ytdChart"></canvas>
      </div>
    </div>

    <div class="section-title">Monthly Pace Table — The Grand Prairie ({month_label(shared_range[0])} – {month_label(shared_range[-1])})</div>
    <div class="chart-card" style="margin-bottom:28px">
      <h3>Net Sales per Month</h3>
      <canvas id="paceLineChart" style="max-height:320px"></canvas>
    </div>
    <div class="table-wrapper" style="margin-bottom:28px;overflow-x:auto">
      <table id="paceTable" class="freeze-col">
        <thead>
          <tr id="paceTableHead"></tr>
        </thead>
        <tbody id="paceTableBody"></tbody>
      </table>
    </div>

    <div class="section-title">Net Sales by Month — The Grand Prairie (GPD + Highlands)</div>
    <div class="chart-card" style="margin-bottom:28px">
      <h3>Homes Sold per Month · Each Year in a Different Color</h3>
      <canvas id="seasonalityChart" style="max-height:340px"></canvas>
    </div>

  </div>

  <!-- GPD TAB -->
  <div id="tab-tgp" class="tab-panel">

    <div class="kpi-grid">
      <div class="kpi tgp"><label>Total Sales (LOP)</label><div class="value">{fmt_num(gpd_lop)}</div><div class="sub">Life of Project</div></div>
      <div class="kpi tgp"><label>2026 Sales / Target</label><div class="value">{gpd_2026_net} / {TARGETS["gpd_annual"]}</div><div class="sub">YTD through {month_names[current_month-1]} · {TARGETS["gpd_monthly"]}/mo target</div></div>
      <div class="kpi tgp"><label>2026 Avg Pace</label><div class="value">{gpd_avg_mo}/mo</div><div class="sub">{ytd_label} avg · target {TARGETS["gpd_monthly"]}/mo</div></div>
      <div class="kpi tgp"><label>Avg Sales Price</label><div class="value">{GPD_AVG_PRICE}</div><div class="sub">Blended LOP avg all lot types</div></div>
    </div>

    <div class="section-title">Monthly Net Sales — GPD ({month_label(d['gpd_months'][0])} – {month_label(d['gpd_months'][-1])})</div>
    <div class="chart-grid single">
      <div class="chart-card">
        <h3>Net Sales by Month</h3>
        <canvas id="tgpMonthChart"></canvas>
      </div>
    </div>

    <div class="chart-grid">
      <div class="chart-card">
        <h3>Sales by Lot Type (LOP)</h3>
        <canvas id="tgpLotChart"></canvas>
      </div>
      <div class="chart-card">
        <h3>Avg. Monthly Pace — 2024 vs 2025 vs 2026 YTD</h3>
        <canvas id="tgpPaceChart"></canvas>
      </div>
    </div>

    <div class="section-title">Builder Summary — GPD</div>
    <div class="table-wrapper">
      <table>
        <thead>
          <tr>
            <th>Lot Type</th>
            <th>Builder</th>
            <th class="right">Avg Pace '24</th>
            <th class="right">Avg Pace '25</th>
            <th class="right">Avg '26 <small style="font-weight:400;opacity:.7">({ytd_label})</small></th>
            <th class="right">Target Pace</th>
            <th class="right">Avg $/SF</th>
            <th class="right">Avg Price</th>
            <th class="right">Total Sales</th>
            <th class="right">% of Total</th>
          </tr>
        </thead>
        <tbody>
{build_gpd_summary()}        </tbody>
      </table>
    </div>

    <div class="section-title">Monthly Net Sales by Lot Size &amp; Builder — GPD</div>
    <div class="chart-card" style="margin-bottom:16px">
      <h3>Net Sales per Month — Stacked by Builder</h3>
      <canvas id="tgpBuilderMonthChart" style="max-height:340px"></canvas>
    </div>
    <div class="table-wrapper">
      <table id="tgpBuilderMonthTable" class="freeze-col">
        <thead><tr id="tgpBuilderMonthHead"></tr></thead>
        <tbody id="tgpBuilderMonthBody"></tbody>
      </table>
    </div>

    <div class="section-title">Net Sales by Month — GPD (Year over Year)</div>
    <div class="chart-card" style="margin-bottom:28px">
      <h3>Homes Sold per Month · Each Year in a Different Color</h3>
      <canvas id="tgpSeasonalityChart" style="max-height:340px"></canvas>
    </div>

  </div>

  <!-- HIGHLANDS TAB -->
  <div id="tab-highlands" class="tab-panel">

    <div class="kpi-grid">
      <div class="kpi hld"><label>Total Sales (LOP)</label><div class="value">{fmt_num(hld_lop)}</div><div class="sub">Life of Project</div></div>
      <div class="kpi hld"><label>2026 Sales / Target</label><div class="value">{hld_2026_net} / {TARGETS["hld_annual"]}</div><div class="sub">YTD through {month_names[current_month-1]} · {TARGETS["hld_monthly"]}/mo target</div></div>
      <div class="kpi hld"><label>2026 Avg Pace</label><div class="value">{hld_avg_mo}/mo</div><div class="sub">{ytd_label} avg · target {TARGETS["hld_monthly"]}/mo</div></div>
      <div class="kpi hld"><label>Avg Sales Price</label><div class="value">{HLD_AVG_PRICE}</div><div class="sub">Blended LOP avg all lot types</div></div>
    </div>

    <div class="section-title">Monthly Net Sales — Highlands ({month_label(d['hld_months'][0])} – {month_label(d['hld_months'][-1])})</div>
    <div class="chart-grid single">
      <div class="chart-card">
        <h3>Net Sales by Month</h3>
        <canvas id="hldMonthChart"></canvas>
      </div>
    </div>

    <div class="chart-grid">
      <div class="chart-card">
        <h3>Sales by Lot Type (LOP)</h3>
        <canvas id="hldLotChart"></canvas>
      </div>
      <div class="chart-card">
        <h3>Avg. Monthly Pace — 2024 vs 2025 vs 2026 YTD</h3>
        <canvas id="hldPaceChart"></canvas>
      </div>
    </div>

    <div class="section-title">Builder Summary — Highlands (Lennar Homes)</div>
    <div class="table-wrapper">
      <table>
        <thead>
          <tr>
            <th>Lot Type</th>
            <th>Builder</th>
            <th class="right">Avg Pace '24</th>
            <th class="right">Avg Pace '25</th>
            <th class="right">Avg '26 <small style="font-weight:400;opacity:.7">({ytd_label})</small></th>
            <th class="right">Avg T12</th>
            <th class="right">Avg Sales Price</th>
            <th class="right">Total Sales</th>
            <th class="right">% of Total</th>
          </tr>
        </thead>
        <tbody>
{build_hld_summary()}        </tbody>
      </table>
    </div>

    <div class="section-title">Monthly Net Sales by Lot Size — Highlands</div>
    <div class="chart-card" style="margin-bottom:16px">
      <h3>Net Sales per Month — Stacked by Lot Size</h3>
      <canvas id="hldBuilderMonthChart" style="max-height:340px"></canvas>
    </div>
    <div class="table-wrapper">
      <table id="hldBuilderMonthTable" class="freeze-col">
        <thead><tr id="hldBuilderMonthHead"></tr></thead>
        <tbody id="hldBuilderMonthBody"></tbody>
      </table>
    </div>

    <div class="section-title">Net Sales by Month — Highlands (Year over Year)</div>
    <div class="chart-card" style="margin-bottom:28px">
      <h3>Homes Sold per Month · Each Year in a Different Color</h3>
      <canvas id="hldSeasonalityChart" style="max-height:340px"></canvas>
    </div>

  </div>

  <!-- WINDROSE GREEN TAB -->
  <div id="tab-wrg" class="tab-panel">

    <div class="kpi-grid">
      <div class="kpi wrg"><label>Total Sales (LOP)</label><div class="value">{fmt_num(wrg_lop)}</div><div class="sub">Life of Project</div></div>
      <div class="kpi wrg"><label>2026 Sales / Target</label><div class="value">{wrg_2026_net} / {TARGETS["wrg_annual"]}</div><div class="sub">YTD through {month_names[current_month-1]} · {TARGETS["wrg_monthly"]}/mo target</div></div>
      <div class="kpi wrg"><label>2026 Avg Pace</label><div class="value">{wrg_avg_mo}/mo</div><div class="sub">{ytd_label} avg · target {TARGETS["wrg_monthly"]}/mo</div></div>
      <div class="kpi wrg"><label>Avg Sales Price</label><div class="value">{WRG_AVG_PRICE}</div><div class="sub">Blended LOP avg all lot types</div></div>
    </div>

    <div class="section-title">Monthly Net Sales — Windrose Green (2022 – {current_year})</div>
    <div class="chart-grid single">
      <div class="chart-card">
        <h3>Net Sales by Month</h3>
        <canvas id="wrgMonthChart"></canvas>
      </div>
    </div>

    <div class="chart-grid">
      <div class="chart-card">
        <h3>Sales by Lot Type (LOP)</h3>
        <canvas id="wrgLotChart"></canvas>
      </div>
      <div class="chart-card">
        <h3>Avg. Monthly Pace — 2024 vs 2025 vs 2026 YTD</h3>
        <canvas id="wrgPaceChart"></canvas>
      </div>
    </div>

    <div class="section-title">Builder Summary — Windrose Green</div>
    <div class="table-wrapper">
      <table>
        <thead>
          <tr>
            <th>Lot Type</th>
            <th>Builder</th>
            <th class="right">Avg Pace '24</th>
            <th class="right">Avg Pace '25</th>
            <th class="right">Avg '26 <small style="font-weight:400;opacity:.7">({ytd_label})</small></th>
            <th class="right">Avg T12</th>
            <th class="right">Avg Sales Price</th>
            <th class="right">Total Sales</th>
            <th class="right">% of Total</th>
          </tr>
        </thead>
        <tbody>
{build_wrg_summary()}        </tbody>
      </table>
    </div>

    <div class="section-title">Monthly Net Sales by Lot Size &amp; Builder — Windrose Green ({month_label(wrg_bm_months[0])} – {month_label(wrg_bm_months[-1])})</div>
    <div class="chart-card" style="margin-bottom:16px">
      <h3>Net Sales per Month — Stacked by Builder</h3>
      <canvas id="wrgBuilderMonthChart" style="max-height:340px"></canvas>
    </div>
    <div class="table-wrapper">
      <table id="wrgBuilderMonthTable" class="freeze-col">
        <thead><tr id="wrgBuilderMonthHead"></tr></thead>
        <tbody id="wrgBuilderMonthBody"></tbody>
      </table>
    </div>

    <div class="section-title">Net Sales by Month — Windrose Green (Year over Year)</div>
    <div class="chart-card" style="margin-bottom:28px">
      <h3>Homes Sold per Month · Each Year in a Different Color</h3>
      <canvas id="wrgSeasonalityChart" style="max-height:340px"></canvas>
    </div>

  </div>

  <!-- 2026 PROJECTIONS -->
  <div id="tab-projections" class="tab-panel">

    <div class="kpi-grid">
      <div class="kpi tgp"><label>TGP 2026 Target</label><div class="value">{TARGETS["gpd_annual"]}</div><div class="sub">{TARGETS["gpd_monthly"]}/mo pace</div></div>
      <div class="kpi hld"><label>Highlands 2026 Target</label><div class="value">{TARGETS["hld_annual"]}</div><div class="sub">{TARGETS["hld_monthly"]}/mo pace</div></div>
      <div class="kpi wrg"><label>WRG 2026 Target</label><div class="value">{TARGETS["wrg_annual"]}</div><div class="sub">{TARGETS["wrg_monthly"]}/mo pace</div></div>
      <div class="kpi total"><label>Combined 2026 Target</label><div class="value">{TARGETS["combined_annual"]}+</div><div class="sub">TGP+Hld+WRG</div></div>
    </div>

    <div class="section-title">2026 Projections by Builder</div>
    <div class="proj-grid">
      <div>
        <div class="section-title" style="margin-top:0">The Grand Prairie</div>
        <div class="table-wrapper">
          <table>
            <thead>
              <tr><th>Lot</th><th>Builder</th><th class="right">Sales Pace</th><th class="right">Annual</th></tr>
            </thead>
            <tbody>
              <tr><td><span class="badge badge-blue">40</span></td><td>David Weekley Homes</td><td class="right">2</td><td class="right">24</td></tr>
              <tr><td><span class="badge badge-blue">40</span></td><td>Perry Homes</td><td class="right">2</td><td class="right">24</td></tr>
              <tr><td><span class="badge badge-green">45</span></td><td>Village Builders</td><td class="right">4</td><td class="right">48</td></tr>
              <tr><td><span class="badge badge-green">45</span></td><td>Westin Homes</td><td class="right">2</td><td class="right">24</td></tr>
              <tr><td><span class="badge badge-orange">50</span></td><td>David Weekley Homes</td><td class="right">3</td><td class="right">36</td></tr>
              <tr><td><span class="badge badge-orange">50</span></td><td>Perry Homes</td><td class="right">3</td><td class="right">36</td></tr>
              <tr><td><span class="badge badge-red">60</span></td><td>David Weekley Homes</td><td class="right">0.75</td><td class="right">9</td></tr>
              <tr><td><span class="badge badge-red">60</span></td><td>Sitterle Homes</td><td class="right">0.75</td><td class="right">9</td></tr>
              <tr><td><span class="badge badge-red">60</span></td><td>Shea Homes</td><td class="right">0.75</td><td class="right">9</td></tr>
              <tr><td><span class="badge badge-red">60</span></td><td>Westin Homes</td><td class="right">0.75</td><td class="right">9</td></tr>
              <tr class="total-row"><td colspan="2"><strong>TGP Total</strong></td><td class="right"><strong>19</strong></td><td class="right"><strong>228</strong></td></tr>
              <tr class="total-row"><td colspan="2">Highlands (40/45/50)</td><td class="right">31</td><td class="right">372</td></tr>
              <tr class="total-row" style="background:#d6e4ff"><td colspan="2"><strong>Combined Total</strong></td><td class="right"><strong>50</strong></td><td class="right"><strong>600</strong></td></tr>
            </tbody>
          </table>
        </div>
        <p style="font-size:.72rem;color:#a0aec0;margin-top:6px">Source: Joe Email 11/05</p>
      </div>
      <div>
        <div class="section-title" style="margin-top:0">Windrose Green</div>
        <div class="table-wrapper">
          <table>
            <thead>
              <tr><th>Builder</th><th>Lots</th><th class="right">Sales Pace</th><th class="right">Annual</th></tr>
            </thead>
            <tbody>
              <tr><td>Lennar Homes</td><td>40/45/50</td><td class="right">14</td><td class="right">168</td></tr>
              <tr><td>CastleRock</td><td>45</td><td class="right">1</td><td class="right">12</td></tr>
              <tr><td>Coventry Homes</td><td>45/50</td><td class="right">3</td><td class="right">36</td></tr>
              <tr class="total-row" style="background:#ffe4e4"><td colspan="2"><strong>WRG Total</strong></td><td class="right"><strong>18</strong></td><td class="right"><strong>216</strong></td></tr>
            </tbody>
          </table>
        </div>

        <div class="section-title">2026 Target vs. Current Pace</div>
        <div class="chart-card" style="margin-top:0">
          <h3>Target vs. Current Pace (units/mo)</h3>
          <canvas id="projChart"></canvas>
        </div>
      </div>
    </div>

  </div>

  <!-- CANCELLATIONS -->
  <div id="tab-cancellations" class="tab-panel">

    <div class="kpi-grid">
      <div class="kpi tgp"><label>GPD Cancellations</label><div class="value">{gpd_canc_lop}</div><div class="sub">Life of Project</div></div>
      <div class="kpi hld"><label>Highlands Cancellations</label><div class="value">{hld_canc_lop}</div><div class="sub">Life of Project</div></div>
      <div class="kpi wrg"><label>WRG Cancellations</label><div class="value">{wrg_canc_lop}</div><div class="sub">Life of Project</div></div>
      <div class="kpi total"><label>Total Cancellations</label><div class="value">{total_canc}</div><div class="sub">All Communities</div></div>
    </div>

    <div class="chart-grid">
      <div class="chart-card">
        <h3>Cancellations by Community</h3>
        <canvas id="cancChart"></canvas>
      </div>
      <div class="chart-card">
        <h3>Cancel Rate vs. Total Gross Sales</h3>
        <canvas id="cancRateChart"></canvas>
      </div>
    </div>

    <div class="section-title">Recent GPD Cancellations</div>
    <div class="table-wrapper">
      <table>
        <thead>
          <tr><th>Address</th><th>Builder</th><th>Lot Type</th><th>Sale Date</th><th>Cancel Date</th><th>Reason</th></tr>
        </thead>
        <tbody>
{build_recent_cancellations()}        </tbody>
      </table>
    </div>

  </div>

  <!-- SALES BY WEEK -->
  <div id="tab-weekly" class="tab-panel">

    <div class="kpi-grid">
      <div class="kpi tgp"><label>GPD 2026 Net Sales</label><div class="value" id="wk-gpd-total">—</div><div class="sub">{ytd_label} 2026</div></div>
      <div class="kpi hld"><label>Highlands 2026 Net Sales</label><div class="value" id="wk-hld-total">—</div><div class="sub">{ytd_label} 2026</div></div>
      <div class="kpi wrg"><label>WRG 2026 Net Sales</label><div class="value" id="wk-wrg-total">—</div><div class="sub">{ytd_label} 2026</div></div>
      <div class="kpi total"><label>Combined 2026 Net</label><div class="value" id="wk-combined-total">—</div><div class="sub">All Communities</div></div>
    </div>

    <div class="section-title">GPD — Weekly Net Sales ({current_year})</div>
    <div class="chart-card" style="margin-bottom:20px">
      <h3>GPD — Net Sales by Week · {ytd_label} {current_year}</h3>
      <canvas id="wkGpdChart"></canvas>
    </div>
    <div class="section-title">GPD — Detail by Lot Size &amp; Builder</div>
    <div class="table-wrapper" style="overflow-x:auto;margin-bottom:28px">
      <table class="freeze-col" id="wkGpdTable"><thead id="wkGpdHead"></thead><tbody id="wkGpdBody"></tbody></table>
    </div>

    <div class="section-title">Highlands — Weekly Net Sales ({current_year})</div>
    <div class="chart-card" style="margin-bottom:20px">
      <h3>Highlands — Net Sales by Week · {ytd_label} {current_year}</h3>
      <canvas id="wkHldChart"></canvas>
    </div>
    <div class="section-title">Highlands — Detail by Lot Size &amp; Builder</div>
    <div class="table-wrapper" style="overflow-x:auto;margin-bottom:36px">
      <table class="freeze-col" id="wkHldTable"><thead id="wkHldHead"></thead><tbody id="wkHldBody"></tbody></table>
    </div>

    <div class="section-title">Windrose Green — Weekly Net Sales ({current_year})</div>
    <div class="chart-card" style="margin-bottom:20px">
      <h3>Windrose Green — Net Sales by Week · {ytd_label} {current_year}</h3>
      <canvas id="wkWrgChart"></canvas>
    </div>
    <div class="section-title">WRG — Detail by Lot Size &amp; Builder</div>
    <div class="table-wrapper" style="overflow-x:auto;margin-bottom:28px">
      <table class="freeze-col" id="wkWrgTable"><thead id="wkWrgHead"></thead><tbody id="wkWrgBody"></tbody></table>
    </div>

  </div>

  <!-- STARTS -->
  <div id="tab-starts" class="tab-panel">

    <div class="kpi-grid" style="grid-template-columns:repeat(2,1fr)">
      <div class="kpi tgp"><label>GPD — Total Inventory</label><div class="value">304</div><div class="sub">Last Drive: Mar 6, 2026</div></div>
      <div class="kpi wrg"><label>WRG — Total Inventory</label><div class="value">311</div><div class="sub">Last Drive: Mar 3, 2026</div></div>
    </div>

    <div class="section-title">GPD — Starts by Month (2026)</div>
    <div class="table-wrapper" style="overflow-x:auto;margin-bottom:36px">
      <table class="freeze-col" id="wkGpdStartsTable"><thead id="wkGpdStartsHead"></thead><tbody id="wkGpdStartsBody"></tbody></table>
    </div>

    <div class="section-title">Windrose Green — Starts by Month (2026)</div>
    <div class="table-wrapper" style="overflow-x:auto;margin-bottom:36px">
      <table class="freeze-col" id="wkWrgStartsTable"><thead id="wkWrgStartsHead"></thead><tbody id="wkWrgStartsBody"></tbody></table>
    </div>

  </div>

</main>

<footer>Sales Dashboard &nbsp;·&nbsp; Data as of {data_date} &nbsp;·&nbsp; The Grand Prairie (GPD + Highlands) · Windrose Green</footer>

<script>
// ── HELPERS ──────────────────────────────────────────────────────────────────
function showTab(name) {{
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('nav button').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  event.target.classList.add('active');
}}

const COLORS = {{
  tgp: '#003552',
  hld: '#38a169',
  wrg: '#e53e3e',
  total: '#d69e2e',
  tgpLight: 'rgba(15,52,96,.15)',
  hldLight: 'rgba(56,161,105,.15)',
  wrgLight: 'rgba(229,62,62,.15)',
}};

Chart.register(ChartDataLabels);
const baseOpts = {{
  responsive: true,
  maintainAspectRatio: true,
  plugins: {{
    legend: {{ position: 'bottom', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }},
    datalabels: {{ display: false }},
  }},
}};

// ── DATA ─────────────────────────────────────────────────────────────────────
const sharedMonths = {js_str_array(shared_labels)};
const tgpShared = {js_array(tgp_shared)};
const hldShared = {js_array(hld_shared)};
const wrgShared = {js_array(wrg_shared)};
const gpShared = sharedMonths.map((_,i) => (tgpShared[i]||0) + (hldShared[i]||0));

const tgpMonths = {js_str_array(tgp_month_labels)};
const tgpMonthData = {js_array(tgp_month_data)};

const hldMonths = {js_str_array(hld_month_labels)};
const hldMonthData = {js_array(hld_month_data)};

const wrgMonths = {js_str_array(wrg_month_labels)};
const wrgMonthData = {js_array(wrg_month_data_arr)};

// ── THE GRAND PRAIRIE TAB CHARTS ─────────────────────────────────────────────
new Chart(document.getElementById('overviewChart'), {{
  type: 'bar',
  data: {{
    labels: sharedMonths,
    datasets: [
      {{ label: 'GPD',       data: tgpShared, backgroundColor: COLORS.tgp, borderRadius: 3 }},
      {{ label: 'Highlands', data: hldShared, backgroundColor: COLORS.hld, borderRadius: 3 }},
    ]
  }},
  options: {{
    ...baseOpts,
    scales: {{
      x: {{
        stacked: true,
        ticks: {{ font: {{ size: 11 }}, color: '#2d3748', maxRotation: 45, minRotation: 45 }}
      }},
      y: {{ stacked: true, ticks: {{ color: '#2d3748' }} }}
    }},
    plugins: {{
      legend: {{ position: 'bottom', labels: {{ boxWidth: 12, font: {{ size: 11 }} }} }},
      datalabels: {{ display: false }},
      tooltip: {{
        mode: 'index', intersect: false,
        callbacks: {{
          footer: (items) => {{
            const total = items.reduce((s, i) => s + i.parsed.y, 0);
            return `Combined: ${{total}}`;
          }}
        }}
      }}
    }}
  }}
}});

new Chart(document.getElementById('mixChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['GPD ({gpd_lop})', 'Highlands ({hld_lop})'],
    datasets: [{{ data: [{gpd_lop}, {hld_lop}], backgroundColor: [COLORS.tgp, COLORS.hld], borderWidth: 2 }}]
  }},
  options: {{
    ...baseOpts,
    plugins: {{
      ...baseOpts.plugins,
      datalabels: {{
        display: true,
        color: '#fff',
        font: {{ weight: 'bold', size: 13 }},
        formatter: (value, ctx) => {{
          const total = ctx.chart.data.datasets[0].data.reduce((a, b) => a + b, 0);
          return ((value / total) * 100).toFixed(1) + '%';
        }},
      }},
    }},
  }},
}});

new Chart(document.getElementById('ytdChart'), {{
  type: 'bar',
  data: {{
    labels: ['GPD', 'Highlands'],
    datasets: [
      {{ label: 'Avg Pace 2024', data: [{gpd_full_pace24}, {hld_full_pace24}], backgroundColor: COLORS.tgp, borderRadius: 3 }},
      {{ label: 'Avg Pace 2025', data: [{gpd_full_pace25}, {hld_full_pace25}], backgroundColor: '#003552', borderRadius: 3 }},
      {{ label: 'Avg Pace 2026 YTD ({ytd_label})', data: [{gpd_avg_mo}, {hld_avg_mo}], backgroundColor: '#f6ad55', borderRadius: 3 }},
    ]
  }},
  options: {{ ...baseOpts, scales: {{ y: {{ beginAtZero: true }} }} }}
}});

// ── TGP CHARTS ────────────────────────────────────────────────────────────────
new Chart(document.getElementById('tgpMonthChart'), {{
  type: 'line',
  data: {{
    labels: tgpMonths,
    datasets: [{{
      label: 'Net Sales',
      data: tgpMonthData,
      borderColor: COLORS.tgp,
      backgroundColor: COLORS.tgpLight,
      fill: true,
      tension: .3,
      pointRadius: 3,
    }}]
  }},
  options: {{ ...baseOpts, scales: {{ x: {{ ticks: {{ font: {{ size: 10 }} }} }}, y: {{ beginAtZero: true }} }} }}
}});

new Chart(document.getElementById('tgpLotChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['40 ft ({gpd_lot_totals.get("40",0)})', '45 ft ({gpd_lot_totals.get("45",0)})', '50 ft ({gpd_lot_totals.get("50",0)})', '60 ft ({gpd_lot_totals.get("60",0)})'],
    datasets: [{{ data: [{gpd_lot_totals.get("40",0)}, {gpd_lot_totals.get("45",0)}, {gpd_lot_totals.get("50",0)}, {gpd_lot_totals.get("60",0)}], backgroundColor: ['#003552','#38a169','#e67e22','#e53e3e'], borderWidth: 2 }}]
  }},
  options: {{
    ...baseOpts,
    plugins: {{
      ...baseOpts.plugins,
      datalabels: {{
        display: true,
        color: '#fff',
        font: {{ weight: 'bold', size: 13 }},
        formatter: (value, ctx) => {{
          const total = ctx.chart.data.datasets[0].data.reduce((a, b) => a + b, 0);
          return ((value / total) * 100).toFixed(1) + '%';
        }},
      }},
    }},
  }},
}});

new Chart(document.getElementById('tgpPaceChart'), {{
  type: 'bar',
  data: {{
    labels: {js_str_array(GPD_PACE_LABELS)},
    datasets: [
      {{ label: 'Avg Pace 2024', data: {js_array(gpd_pace24)}, backgroundColor: COLORS.tgp, borderRadius: 3 }},
      {{ label: 'Avg Pace 2025', data: {js_array(gpd_pace25)}, backgroundColor: '#003552', borderRadius: 3 }},
      {{ label: 'Avg Pace 2026 YTD ({ytd_label})', data: {js_array(gpd_pace26)}, backgroundColor: '#f6ad55', borderRadius: 3 }},
    ]
  }},
  options: {{ ...baseOpts, scales: {{ x: {{ ticks: {{ font: {{ size: 10 }} }} }}, y: {{ beginAtZero: true }} }} }}
}});

// ── HIGHLANDS CHARTS ──────────────────────────────────────────────────────────
new Chart(document.getElementById('hldMonthChart'), {{
  type: 'line',
  data: {{
    labels: hldMonths,
    datasets: [{{
      label: 'Net Sales',
      data: hldMonthData,
      borderColor: COLORS.hld,
      backgroundColor: COLORS.hldLight,
      fill: true,
      tension: .3,
      pointRadius: 3,
    }}]
  }},
  options: {{
    ...baseOpts,
    scales: {{
      x: {{ ticks: {{ font: {{ size: 10 }}, color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }},
      y: {{ beginAtZero: true, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }}
    }},
    plugins: {{
      ...baseOpts.plugins,
      legend: {{ position: 'bottom', labels: {{ boxWidth: 12, font: {{ size: 11 }}, color: '#000' }} }},
      datalabels: {{ display: false }}
    }}
  }}
}});

new Chart(document.getElementById('hldLotChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['40 ft ({hld_lot_totals.get("40",0)})', '45 ft ({hld_lot_totals.get("45",0)})', '50 ft ({hld_lot_totals.get("50",0)})'],
    datasets: [{{ data: [{hld_lot_totals.get("40",0)}, {hld_lot_totals.get("45",0)}, {hld_lot_totals.get("50",0)}], backgroundColor: ['#003552','#38a169','#e67e22'], borderWidth: 2 }}]
  }},
  options: {{
    ...baseOpts,
    plugins: {{
      ...baseOpts.plugins,
      datalabels: {{
        display: true,
        color: '#fff',
        font: {{ weight: 'bold', size: 13 }},
        formatter: (value, ctx) => {{
          const total = ctx.chart.data.datasets[0].data.reduce((a, b) => a + b, 0);
          return ((value / total) * 100).toFixed(1) + '%';
        }},
      }},
    }},
  }},
}});

new Chart(document.getElementById('hldPaceChart'), {{
  type: 'bar',
  data: {{
    labels: ['Lennar 40', 'Lennar 45', 'Lennar 50'],
    datasets: [
      {{ label: 'Avg Pace 2024', data: {js_array(hld_pace24)}, backgroundColor: COLORS.hld, borderRadius: 4 }},
      {{ label: 'Avg Pace 2025', data: {js_array(hld_pace25)}, backgroundColor: '#68d391', borderRadius: 4 }},
      {{ label: 'Avg Pace 2026 YTD ({ytd_label})', data: {js_array(hld_pace26)}, backgroundColor: '#f6ad55', borderRadius: 4 }},
    ]
  }},
  options: {{ ...baseOpts, scales: {{ y: {{ beginAtZero: true }} }} }}
}});

// ── WRG CHARTS ────────────────────────────────────────────────────────────────
new Chart(document.getElementById('wrgMonthChart'), {{
  type: 'line',
  data: {{
    labels: wrgMonths,
    datasets: [{{
      label: 'Net Sales',
      data: wrgMonthData,
      borderColor: COLORS.wrg,
      backgroundColor: COLORS.wrgLight,
      fill: true,
      tension: .3,
      pointRadius: 3,
    }}]
  }},
  options: {{ ...baseOpts, scales: {{ x: {{ ticks: {{ font: {{ size: 10 }} }} }}, y: {{ beginAtZero: true }} }} }}
}});

new Chart(document.getElementById('wrgLotChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['40 ft ({wrg_lot_totals.get("40",0)})', '45 ft ({wrg_lot_totals.get("45",0)})', '50 ft ({wrg_lot_totals.get("50",0)})'],
    datasets: [{{ data: [{wrg_lot_totals.get("40",0)}, {wrg_lot_totals.get("45",0)}, {wrg_lot_totals.get("50",0)}], backgroundColor: ['#003552','#38a169','#e67e22'], borderWidth: 2 }}]
  }},
  options: {{
    ...baseOpts,
    plugins: {{
      ...baseOpts.plugins,
      datalabels: {{
        display: true,
        color: '#fff',
        font: {{ weight: 'bold', size: 13 }},
        formatter: (value, ctx) => {{
          const total = ctx.chart.data.datasets[0].data.reduce((a, b) => a + b, 0);
          return ((value / total) * 100).toFixed(1) + '%';
        }},
      }},
    }},
  }},
}});

new Chart(document.getElementById('wrgPaceChart'), {{
  type: 'bar',
  data: {{
    labels: {js_str_array(WRG_PACE_LABELS)},
    datasets: [
      {{ label: 'Avg Pace 2024', data: {js_array(wrg_pace24)}, backgroundColor: COLORS.wrg, borderRadius: 3 }},
      {{ label: 'Avg Pace 2025', data: {js_array(wrg_pace25)}, backgroundColor: '#fc8181', borderRadius: 3 }},
      {{ label: 'Avg Pace 2026 YTD ({ytd_label})', data: {js_array(wrg_pace26)}, backgroundColor: '#f6ad55', borderRadius: 3 }},
    ]
  }},
  options: {{ ...baseOpts, scales: {{ x: {{ ticks: {{ font: {{ size: 9 }} }} }}, y: {{ beginAtZero: true }} }} }}
}});

// ── MONTHLY PACE LINE CHART + TABLE ──────────────────────────────────────────
const paceMonths = {js_str_array(shared_labels)};
const paceTGP  = {js_array(tgp_shared)};
const paceHLD  = {js_array(hld_shared)};
const paceWRG  = {js_array(wrg_shared)};
const paceGP    = paceMonths.map((_,i) => (paceTGP[i]||0)+(paceHLD[i]||0));
const paceTotal = paceMonths.map((_,i) => (paceTGP[i]||0)+(paceHLD[i]||0)+(paceWRG[i]||0));

new Chart(document.getElementById('paceLineChart'), {{
  type: 'line',
  data: {{
    labels: paceMonths,
    datasets: [
      {{ label: 'GPD',              data: paceTGP,  borderColor: COLORS.tgp,   backgroundColor: 'transparent', tension:.3, pointRadius:3, borderWidth:2 }},
      {{ label: 'Highlands',        data: paceHLD,  borderColor: COLORS.hld,   backgroundColor: 'transparent', tension:.3, pointRadius:3, borderWidth:2 }},
      {{ label: 'Combined GP',      data: paceGP,   borderColor: COLORS.total, backgroundColor: 'transparent', tension:.3, pointRadius:3, borderWidth:2.5, borderDash:[5,3] }},
    ]
  }},
  options: {{
    ...baseOpts,
    scales: {{
      x: {{ ticks: {{ font: {{ size: 10 }}, maxRotation: 45 }} }},
      y: {{ beginAtZero: true, title: {{ display: true, text: 'Net Sales / Month' }} }}
    }}
  }}
}});

(function buildPaceTable() {{
  const head = document.getElementById('paceTableHead');
  const body = document.getElementById('paceTableBody');
  const rows = [
    {{ label: 'GPD',         data: paceTGP, cls: 'tgp' }},
    {{ label: 'Highlands',   data: paceHLD, cls: 'hld' }},
    {{ label: 'Combined GP', data: paceGP,  cls: 'total' }},
  ];
  let hHtml = '<th>Community</th>';
  paceMonths.forEach(m => hHtml += `<th class="right" style="font-size:.72rem;white-space:nowrap">${{m}}</th>`);
  head.innerHTML = hHtml;
  rows.forEach(r => {{
    let cells = `<td><span class="badge badge-${{r.cls === 'tgp' ? 'blue' : r.cls === 'hld' ? 'green' : r.cls === 'wrg' ? 'red' : 'orange'}}" style="white-space:nowrap">${{r.label}}</span></td>`;
    r.data.forEach(v => {{
      const style = v === 0 ? 'color:#cbd5e0' : v >= 50 ? 'font-weight:700;color:#276749' : v >= 20 ? 'color:#003552' : '';
      cells += `<td class="right" style="font-size:.8rem;${{style}}">${{v > 0 ? v : '—'}}</td>`;
    }});
    const trClass = r.cls === 'total' ? 'class="total-row"' : '';
    body.innerHTML += `<tr ${{trClass}}>${{cells}}</tr>`;
  }});
}})();

// ── SEASONALITY CHART — year-over-year lines ─────────────────────────────────
(function() {{
  const moLabels = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  const gpd = {yoy_js(gpd_yoy)};
  const hld = {yoy_js(hld_yoy)};
  const years = [{','.join(str(y) for y in sorted(combined_yoy.keys()))}];
  const combined = {{}};
  years.forEach(yr => {{
    combined[yr] = moLabels.map((_,i) => {{
      const g = gpd[yr] ? gpd[yr][i] : null;
      const h = hld[yr] ? hld[yr][i] : null;
      if (g === null && h === null) return null;
      return (g || 0) + (h || 0);
    }});
  }});

  const palette = {{
    2023: {{ border: '#a0aec0', bg: 'rgba(160,174,192,.15)' }},
    2024: {{ border: '#003552', bg: 'rgba(0,53,82,.12)' }},
    2025: {{ border: '#48bb78', bg: 'rgba(72,187,120,.12)' }},
    2026: {{ border: '#fc8181', bg: 'rgba(252,129,129,.15)' }},
  }};

  new Chart(document.getElementById('seasonalityChart'), {{
    type: 'line',
    data: {{
      labels: moLabels,
      datasets: years.map(yr => ({{
        label: String(yr),
        data: combined[yr],
        borderColor: palette[yr].border,
        backgroundColor: palette[yr].bg,
        borderWidth: yr === {current_year} ? 3 : 2,
        pointRadius: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 9 : 5,
        pointHoverRadius: 7,
        pointBackgroundColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#e53e3e' : palette[yr].border,
        pointBorderColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#c53030' : palette[yr].border,
        pointStyle: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 'triangle' : 'circle',
        segment: {{
          borderColor: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? '#e53e3e' : undefined,
          borderDash: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? [4,3] : undefined,
        }},
        fill: false,
        tension: 0.3,
        spanGaps: false,
      }}))
    }},
    options: {{
      ...baseOpts,
      scales: {{
        y: {{
          beginAtZero: true,
          title: {{ display: true, text: 'Net Sales (homes)', color: '#000' }},
          ticks: {{ color: '#000' }},
          grid: {{ color: 'rgba(0,0,0,.08)' }}
        }},
        x: {{
          title: {{ display: true, text: 'Month', color: '#000' }},
          ticks: {{ color: '#000' }},
          grid: {{ color: 'rgba(0,0,0,.08)' }}
        }}
      }},
      plugins: {{
        legend: {{
          position: 'top',
          labels: {{ boxWidth: 14, font: {{ size: 12 }}, color: '#000' }}
        }},
        tooltip: {{
          mode: 'index',
          intersect: false,
          callbacks: {{
            label: ctx => `${{ctx.dataset.label}}: ${{ctx.parsed.y !== null ? ctx.parsed.y + ' homes' : '—'}}`
          }}
        }}
      }}
    }}
  }});
}})();

// ── SEASONALITY — GPD only ─────────────────────────────────────────────────────
(function() {{
  const moLabels = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  const yearData = {yoy_js(gpd_yoy)};
  const palette = {{
    2023: {{ border: '#a0aec0' }}, 2024: {{ border: '#003552' }},
    2025: {{ border: '#48bb78' }}, 2026: {{ border: '#fc8181' }},
  }};
  new Chart(document.getElementById('tgpSeasonalityChart'), {{
    type: 'line',
    data: {{
      labels: moLabels,
      datasets: Object.keys(yearData).map(yr => ({{
        label: yr,
        data: yearData[yr],
        borderColor: palette[yr].border,
        backgroundColor: 'transparent',
        borderWidth: yr == {current_year} ? 3 : 2,
        pointRadius: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 9 : 5,
        pointHoverRadius: 7,
        pointBackgroundColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#e53e3e' : palette[yr].border,
        pointBorderColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#c53030' : palette[yr].border,
        pointStyle: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 'triangle' : 'circle',
        segment: {{
          borderColor: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? '#e53e3e' : undefined,
          borderDash: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? [4,3] : undefined,
        }},
        fill: false, tension: 0.3, spanGaps: false,
      }}))
    }},
    options: {{
      ...baseOpts,
      scales: {{
        y: {{ title: {{ display: true, text: 'Net Sales (homes)', color: '#000' }}, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }},
        x: {{ title: {{ display: true, text: 'Month', color: '#000' }}, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }}
      }},
      plugins: {{
        legend: {{ position: 'top', labels: {{ boxWidth: 14, font: {{ size: 12 }}, color: '#000' }} }},
        tooltip: {{ mode: 'index', intersect: false, callbacks: {{ label: ctx => `${{ctx.dataset.label}}: ${{ctx.parsed.y !== null ? ctx.parsed.y + ' homes' : '—'}}` }} }}
      }}
    }}
  }});
}})();

// ── SEASONALITY — Highlands only ───────────────────────────────────────────────
(function() {{
  const moLabels = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  const yearData = {yoy_js(hld_yoy)};
  const palette = {{
    2024: {{ border: '#003552' }}, 2025: {{ border: '#48bb78' }}, 2026: {{ border: '#fc8181' }},
  }};
  new Chart(document.getElementById('hldSeasonalityChart'), {{
    type: 'line',
    data: {{
      labels: moLabels,
      datasets: Object.keys(yearData).map(yr => ({{
        label: yr,
        data: yearData[yr],
        borderColor: palette[yr].border,
        backgroundColor: 'transparent',
        borderWidth: yr == {current_year} ? 3 : 2,
        pointRadius: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 9 : 5,
        pointHoverRadius: 7,
        pointBackgroundColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#e53e3e' : palette[yr].border,
        pointBorderColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#c53030' : palette[yr].border,
        pointStyle: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 'triangle' : 'circle',
        segment: {{
          borderColor: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? '#e53e3e' : undefined,
          borderDash: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? [4,3] : undefined,
        }},
        fill: false, tension: 0.3, spanGaps: false,
      }}))
    }},
    options: {{
      ...baseOpts,
      scales: {{
        y: {{ title: {{ display: true, text: 'Net Sales (homes)', color: '#000' }}, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }},
        x: {{ title: {{ display: true, text: 'Month', color: '#000' }}, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }}
      }},
      plugins: {{
        legend: {{ position: 'top', labels: {{ boxWidth: 14, font: {{ size: 12 }}, color: '#000' }} }},
        tooltip: {{ mode: 'index', intersect: false, callbacks: {{ label: ctx => `${{ctx.dataset.label}}: ${{ctx.parsed.y !== null ? ctx.parsed.y + ' homes' : '—'}}` }} }}
      }}
    }}
  }});
}})();

// ── SEASONALITY — Windrose Green only ─────────────────────────────────────────
(function() {{
  const moLabels = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  const yearData = {yoy_js(wrg_yoy)};
  const palette = {{
    2024: {{ border: '#003552' }}, 2025: {{ border: '#48bb78' }}, 2026: {{ border: '#fc8181' }},
  }};
  new Chart(document.getElementById('wrgSeasonalityChart'), {{
    type: 'line',
    data: {{
      labels: moLabels,
      datasets: Object.keys(yearData).map(yr => ({{
        label: yr,
        data: yearData[yr],
        borderColor: palette[yr].border,
        backgroundColor: 'transparent',
        borderWidth: yr == {current_year} ? 3 : 2,
        pointRadius: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 9 : 5,
        pointHoverRadius: 7,
        pointBackgroundColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#e53e3e' : palette[yr].border,
        pointBorderColor: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? '#c53030' : palette[yr].border,
        pointStyle: (ctx) => (ctx.parsed?.y ?? 0) < 0 ? 'triangle' : 'circle',
        segment: {{
          borderColor: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? '#e53e3e' : undefined,
          borderDash: (ctx) => (ctx.p0.parsed.y < 0 || ctx.p1.parsed.y < 0) ? [4,3] : undefined,
        }},
        fill: false, tension: 0.3, spanGaps: false,
      }}))
    }},
    options: {{
      ...baseOpts,
      scales: {{
        y: {{ title: {{ display: true, text: 'Net Sales (homes)', color: '#000' }}, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }},
        x: {{ title: {{ display: true, text: 'Month', color: '#000' }}, ticks: {{ color: '#000' }}, grid: {{ color: 'rgba(0,0,0,.08)' }} }}
      }},
      plugins: {{
        legend: {{ position: 'top', labels: {{ boxWidth: 14, font: {{ size: 12 }}, color: '#000' }} }},
        tooltip: {{ mode: 'index', intersect: false, callbacks: {{ label: ctx => `${{ctx.dataset.label}}: ${{ctx.parsed.y !== null ? ctx.parsed.y + ' homes' : '—'}}` }} }}
      }}
    }}
  }});
}})();

// ── PROJECTIONS CHART ─────────────────────────────────────────────────────────
new Chart(document.getElementById('projChart'), {{
  type: 'bar',
  data: {{
    labels: ['GPD', 'Highlands', 'WRG'],
    datasets: [
      {{ label: '2026 Target (mo)', data: [{TARGETS["gpd_monthly"]}, {TARGETS["hld_monthly"]}, {TARGETS["wrg_monthly"]}], backgroundColor: '#003552', borderRadius: 4 }},
      {{ label: 'Current 2026 Pace', data: [{gpd_avg_mo}, {hld_avg_mo}, {wrg_avg_mo}], backgroundColor: '#fc8181', borderRadius: 4 }},
    ]
  }},
  options: {{ ...baseOpts, scales: {{ y: {{ beginAtZero: true }} }} }}
}});

// ── CANCELLATION CHARTS ───────────────────────────────────────────────────────
new Chart(document.getElementById('cancChart'), {{
  type: 'bar',
  data: {{
    labels: ['GPD', 'Highlands', 'WRG'],
    datasets: [{{ label: 'Cancellations', data: [{gpd_canc_lop}, {hld_canc_lop}, {wrg_canc_lop}], backgroundColor: [COLORS.tgp, COLORS.hld, COLORS.wrg], borderRadius: 6 }}]
  }},
  options: {{
    ...baseOpts,
    plugins: {{
      legend: {{ display: false }},
      datalabels: {{
        display: true,
        anchor: 'end',
        align: 'end',
        color: '#2d3748',
        font: {{ weight: 'bold', size: 12 }},
        formatter: (value) => {{
          const total = {total_canc};
          return value + ' (' + ((value / total) * 100).toFixed(1) + '%)';
        }}
      }}
    }},
    scales: {{ y: {{ beginAtZero: true, max: {int(max(gpd_canc_lop, hld_canc_lop, wrg_canc_lop) * 1.2)} }} }},
    layout: {{ padding: {{ top: 24 }} }}
  }}
}});

new Chart(document.getElementById('cancRateChart'), {{
  type: 'bar',
  data: {{
    labels: ['GPD', 'Highlands', 'WRG'],
    datasets: [
      {{ label: 'Gross Sales', data: [{gpd_gross_lop}, {hld_gross_lop}, {wrg_gross_lop}], backgroundColor: ['rgba(15,52,96,.4)','rgba(56,161,105,.4)','rgba(229,62,62,.4)'], borderRadius: 4 }},
      {{ label: 'Cancellations', data: [{gpd_canc_lop}, {hld_canc_lop}, {wrg_canc_lop}], backgroundColor: [COLORS.tgp, COLORS.hld, COLORS.wrg], borderRadius: 4 }},
    ]
  }},
  options: {{
    ...baseOpts,
    plugins: {{
      ...baseOpts.plugins,
      datalabels: {{
        display: (ctx) => ctx.datasetIndex === 1,
        anchor: 'end',
        align: 'end',
        color: '#2d3748',
        font: {{ weight: 'bold', size: 12 }},
        formatter: (value, ctx) => {{
          const grossSales = [{gpd_gross_lop}, {hld_gross_lop}, {wrg_gross_lop}];
          const rate = ((value / grossSales[ctx.dataIndex]) * 100).toFixed(1);
          return rate + '% canc.';
        }}
      }}
    }},
    scales: {{ y: {{ beginAtZero: true }} }}
  }}
}});

// ── MONTHLY BY LOT SIZE & BUILDER ────────────────────────────────────────────
const BP = [
  '#003552','#38a169','#e53e3e','#d69e2e','#805ad5',
  '#0987a0','#c05621','#b83280','#2f855a','#718096'
];

function buildBuilderMonthChart(canvasId, months, seriesMap) {{
  const labels = Object.keys(seriesMap);
  const datasets = labels.map((lbl, i) => ({{
    label: lbl,
    data: seriesMap[lbl],
    backgroundColor: BP[i % BP.length],
    borderRadius: 2,
    borderSkipped: false,
  }}));
  new Chart(document.getElementById(canvasId), {{
    type: 'bar',
    data: {{ labels: months, datasets }},
    options: {{
      ...baseOpts,
      scales: {{
        x: {{ stacked: true, ticks: {{ font: {{ size: 9 }}, maxRotation: 45 }} }},
        y: {{ stacked: true, beginAtZero: true, title: {{ display: true, text: 'Net Sales' }} }}
      }},
      plugins: {{
        legend: {{ position: 'bottom', labels: {{ boxWidth: 12, font: {{ size: 10 }} }} }},
        tooltip: {{ mode: 'index', intersect: false }},
        datalabels: {{ display: false }}
      }}
    }}
  }});
}}

function buildBuilderMonthTable(headId, bodyId, months, seriesMap) {{
  const head = document.getElementById(headId);
  const body = document.getElementById(bodyId);
  const builderNames = {{
    'DWH':'David Weekley Homes','Perry':'Perry Homes','Village':'Village Builders',
    'Westin':'Westin Homes','K.Hov':'K. Hovnanian','Lennar':'Lennar Homes',
    'CastleRock':'CastleRock Homes','Coventry':'Coventry Homes',
    'Sitterle':'Sitterle Homes','Shea':'Shea Homes',
  }};
  const lotColors = {{'40':'#003552','45':'#48bb78','50':'#ed8936','60':'#e53e3e'}};

  let hHtml = '<th style="white-space:nowrap;min-width:64px;text-align:center">Lot Type</th>';
  hHtml += '<th style="white-space:nowrap;min-width:150px">Builder</th>';
  months.forEach(m => hHtml += `<th class="right" style="font-size:.7rem;white-space:nowrap;padding:8px 8px">${{m}}</th>`);
  hHtml += `<th class="right" style="font-size:.7rem;white-space:nowrap;padding:8px 8px">Total</th>`;
  head.innerHTML = hHtml;

  const groups = {{}};
  Object.entries(seriesMap).forEach(([lbl, data]) => {{
    const parts = lbl.split(' – ');
    const lotSize = parts[0].replace('ft','').trim();
    const builderShort = parts[1] ? parts[1].trim() : lbl;
    const builderFull = builderNames[builderShort] || builderShort;
    if (!groups[lotSize]) groups[lotSize] = [];
    groups[lotSize].push({{ builder: builderFull, data }});
  }});

  const grandMonthTotals = months.map((_, mi) =>
    Object.values(seriesMap).reduce((s, arr) => s + (arr[mi] || 0), 0)
  );
  const grandTotal = grandMonthTotals.reduce((s, v) => s + v, 0);

  const cellStyle = (v) => {{
    if (v === 0 || v == null) return `<td class="right" style="font-size:.78rem;color:#cbd5e0">—</td>`;
    const clr = v < 0 ? 'color:#c53030;font-weight:700' : 'color:#2d3748';
    return `<td class="right" style="font-size:.78rem;${{clr}}">${{v}}</td>`;
  }};

  let bHtml = '';
  Object.entries(groups).forEach(([lotSize, rows]) => {{
    const color = lotColors[lotSize] || '#a0aec0';
    const groupMonthTotals = months.map((_, mi) => rows.reduce((s, r) => s + (r.data[mi] || 0), 0));
    const groupTotal = groupMonthTotals.reduce((s, v) => s + v, 0);
    rows.forEach((row, ri) => {{
      const rowTotal = row.data.reduce((s, v) => s + (v > 0 ? v : 0), 0);
      const cells = row.data.map(v => cellStyle(v)).join('');
      const lotCell = ri === 0
        ? `<td style="text-align:center;vertical-align:middle">
             <span style="display:inline-block;background:${{color}};color:#fff;border-radius:4px;padding:2px 9px;font-size:.75rem;font-weight:700">${{lotSize}}</span>
           </td>`
        : `<td></td>`;
      bHtml += `<tr>${{lotCell}}
        <td style="white-space:nowrap;font-size:.78rem;padding-left:6px">${{row.builder}}</td>
        ${{cells}}<td class="right" style="font-size:.78rem;font-weight:700;color:#003552">${{rowTotal}}</td>
      </tr>`;
    }});
    const grpCells = groupMonthTotals.map(v => `<td class="right" style="font-size:.78rem;font-weight:700">${{v || '—'}}</td>`).join('');
    bHtml += `<tr class="total-row">
      <td colspan="2" style="font-size:.78rem;font-weight:700;padding-left:8px">${{lotSize}} Total</td>
      ${{grpCells}}<td class="right" style="font-size:.78rem;font-weight:700">${{groupTotal}}</td>
    </tr>`;
  }});

  const grandCells = grandMonthTotals.map(v => `<td class="right" style="font-size:.8rem;font-weight:800;color:#003552">${{v || '—'}}</td>`).join('');
  bHtml += `<tr class="total-row">
    <td colspan="2" style="font-size:.8rem;font-weight:800;padding-left:8px;color:#003552">Grand Total</td>
    ${{grandCells}}<td class="right" style="font-size:.8rem;font-weight:800;color:#003552">${{grandTotal}}</td>
  </tr>`;
  body.innerHTML = bHtml;
}}

// ── Builder-by-month data ─────────────────────────────────────────────────
const tgpBMMonths = {js_str_array(tgp_month_labels)};
const tgpBMData = {tgp_bm_js};
buildBuilderMonthChart('tgpBuilderMonthChart', tgpBMMonths, tgpBMData);
buildBuilderMonthTable('tgpBuilderMonthHead', 'tgpBuilderMonthBody', tgpBMMonths, tgpBMData);

const hldBMMonths = {js_str_array(hld_month_labels)};
const hldBMData = {hld_bm_js};
buildBuilderMonthChart('hldBuilderMonthChart', hldBMMonths, hldBMData);
buildBuilderMonthTable('hldBuilderMonthHead', 'hldBuilderMonthBody', hldBMMonths, hldBMData);

const wrgBMMonths = {js_str_array([month_label(m) for m in wrg_bm_months])};
const wrgBMData = {wrg_bm_js};
buildBuilderMonthChart('wrgBuilderMonthChart', wrgBMMonths, wrgBMData);
buildBuilderMonthTable('wrgBuilderMonthHead', 'wrgBuilderMonthBody', wrgBMMonths, wrgBMData);

// ── SALES BY WEEK (2026 ONLY) ─────────────────────────────────────────────────
const gpdWkLabels = {js_str_array(gpd_wk['labels'])};
const hldWkLabels = {js_str_array(hld_wk['labels'])};
const wrgWkLabels = {js_str_array(wrg_wk['labels'])};

const gpdWkGross={js_array(gpd_wk['gross'])}, gpdWkCanc={js_array(gpd_wk['cancel'])};
const hldWkGross={js_array(hld_wk['gross'])}, hldWkCanc={js_array(hld_wk['cancel'])};
const wrgWkGross={js_array(wrg_wk['gross'])}, wrgWkCanc={js_array(wrg_wk['cancel'])};
const gpdWkNet=gpdWkGross.map((v,i)=>v+gpdWkCanc[i]);
const hldWkNet=hldWkGross.map((v,i)=>v+hldWkCanc[i]);
const wrgWkNet=wrgWkGross.map((v,i)=>v+wrgWkCanc[i]);

const gpdWkRows = {gpd_wk_rows_js};
const hldWkRows = {hld_wk_rows_js};
const wrgWkRows = {wrg_wk_rows_js};

// Starts data (manually maintained — update when new drive data available)
const gpdStartsMonths = ['Jan 26','Feb 26 (1)','Feb 26 (2)','Mar 26'];
const gpdStartsDates  = ['01/21/26','02/02/26','02/19/26','03/06/26'];
const gpdStartsRows = {d.get('gpd_starts_js', '[]')};

const wrgStartsMonths = ['Jan 26','Feb 26','Mar 26'];
const wrgStartsDates  = ['01/20/26','02/13/26','03/03/26'];
const wrgStartsRows = {d.get('wrg_starts_js', '[]')};

// ── CHART BUILDER ─────────────────────────────────────────────────────────────
function rollingAvg(arr, n) {{
  return arr.map((_, i) => {{
    const slice = arr.slice(Math.max(0, i - n + 1), i + 1);
    return +(slice.reduce((a, b) => a + b, 0) / slice.length).toFixed(2);
  }});
}}

function buildWeeklyChart(canvasId, labels, gross, canc, net, color, colorLight) {{
  const avg4 = rollingAvg(net, 4);
  new Chart(document.getElementById(canvasId), {{
    type: 'bar',
    data: {{
      labels,
      datasets: [
        {{
          type: 'bar', label: 'Gross Sales', data: gross,
          backgroundColor: colorLight, borderColor: color, borderWidth: 1, borderRadius: 4, order: 3,
        }},
        {{
          type: 'bar', label: 'Cancellations', data: canc,
          backgroundColor: 'rgba(229,62,62,0.25)', borderColor: '#e53e3e', borderWidth: 1, borderRadius: 4, order: 3,
        }},
        {{
          type: 'line', label: 'Net Sales', data: net,
          borderColor: color, borderWidth: 2.5,
          pointRadius: net.map(v => v < 0 ? 6 : 4),
          pointBackgroundColor: net.map(v => v < 0 ? '#e53e3e' : color),
          pointStyle: net.map(v => v < 0 ? 'triangle' : 'circle'),
          tension: 0.3, fill: false, order: 2,
        }},
        {{
          type: 'line', label: '4-Wk Avg', data: avg4,
          borderColor: '#d69e2e', borderWidth: 2, borderDash: [5, 3],
          pointRadius: 0, tension: 0.4, fill: false, order: 1,
        }}
      ]
    }},
    options: {{
      ...baseOpts,
      plugins: {{
        ...baseOpts.plugins,
        legend: {{ display: true, labels: {{ color: '#000', font: {{ size: 11 }}, boxWidth: 12 }} }},
        datalabels: {{ display: false }}
      }},
      scales: {{
        x: {{ grid: {{ color: 'rgba(0,0,0,.06)' }}, ticks: {{ color: '#000', maxRotation: 45, font: {{ size: 10 }} }} }},
        y: {{ grid: {{ color: 'rgba(0,0,0,.08)' }}, ticks: {{ color: '#000' }}, title: {{ display: true, text: 'Homes', color: '#000', font: {{ size: 11 }} }} }}
      }}
    }}
  }});
}}

function buildWeeklyDetailTable(headId, bodyId, labels, rows) {{
  const thead = document.getElementById(headId);
  const tbody = document.getElementById(bodyId);
  const typeLabels = {{ g:'Gross', c:'Cancellations', n:'Net Sales' }};
  const typeColors = {{ g:'#003552', c:'#c53030', n:'#276749' }};
  const H = 'background:#003552;color:#fff';
  let th = `<tr>`
    + `<th style="position:sticky;left:0;z-index:3;${{H}};min-width:50px">Lot</th>`
    + `<th style="position:sticky;left:50px;z-index:3;${{H}};min-width:140px">Builder</th>`
    + `<th style="position:sticky;left:190px;z-index:3;${{H}};min-width:100px">Type</th>`
    + `<th style="${{H}};min-width:60px">LOP Total</th>`;
  labels.forEach(l => th += `<th style="${{H}};min-width:52px">${{l}}</th>`);
  th += '</tr>';
  thead.innerHTML = th;
  const lots = [...new Set(rows.map(r => r.lot))];
  let html = '';
  lots.forEach(lot => {{
    const lotRows = rows.filter(r => r.lot === lot);
    const builders = [...new Set(lotRows.map(r => r.bld))];
    let lotFirst = true;
    builders.forEach(bld => {{
      ['g','c','n'].forEach(type => {{
        const row = lotRows.find(r => r.bld === bld);
        if (!row) return;
        const vals = row[type];
        const lopTotal = type === 'g' ? row.tg : (type === 'c' ? row.tc : row.tn);
        const clr = typeColors[type];
        let tr = `<tr style="border-bottom:1px solid #e2e8f0">`;
        if (type === 'g') {{
          tr += `<td style="position:sticky;left:0;z-index:2;background:#fff;font-weight:700;text-align:center">${{lotFirst && type==='g' ? lot+'ft' : ''}}</td>`;
          tr += `<td style="position:sticky;left:50px;z-index:2;background:#fff;font-size:.82rem">${{bld}}</td>`;
        }} else {{
          tr += `<td style="position:sticky;left:0;z-index:2;background:#fff"></td>`;
          tr += `<td style="position:sticky;left:50px;z-index:2;background:#fff"></td>`;
        }}
        tr += `<td style="position:sticky;left:190px;z-index:2;background:#fff;color:${{clr}};font-weight:600;font-size:.8rem">${{typeLabels[type]}}</td>`;
        tr += `<td style="text-align:right;color:${{lopTotal<0?'#c53030':'#2d3748'}};font-weight:600">${{lopTotal}}</td>`;
        vals.forEach(v => {{
          const c = v < 0 ? 'color:#c53030;font-weight:700' : (v > 0 ? 'color:#276749' : 'color:#a0aec0');
          tr += `<td style="text-align:right;${{c}}">${{v !== 0 ? v : ''}}</td>`;
        }});
        tr += '</tr>';
        html += tr;
        if (type === 'g') lotFirst = false;
      }});
      html += `<tr style="height:4px;background:#f7fafc"><td colspan="${{4 + labels.length}}"></td></tr>`;
    }});
    const lotNet = rows.filter(r => r.lot === lot).reduce((sum, r) => sum + r.tn, 0);
    html += `<tr style="background:#ebf8ff;font-weight:700;border-top:2px solid #bee3f8">
      <td style="position:sticky;left:0;z-index:2;background:#ebf8ff;text-align:center">${{lot}}ft</td>
      <td style="position:sticky;left:50px;z-index:2;background:#ebf8ff;font-size:.82rem">Subtotal</td>
      <td style="position:sticky;left:190px;z-index:2;background:#ebf8ff;color:#276749;font-weight:700">Net</td>
      <td style="text-align:right;color:${{lotNet<0?'#c53030':'#276749'}}">${{lotNet}}</td>`;
    const weekSums = labels.map((_,i) => rows.filter(r=>r.lot===lot).reduce((s,r)=>s+r.n[i],0));
    weekSums.forEach(v => {{
      html += `<td style="text-align:right;color:${{v<0?'#c53030':v>0?'#276749':'#a0aec0'}};font-weight:700">${{v||''}}</td>`;
    }});
    html += '</tr>';
  }});
  const grandNet = rows.reduce((s,r)=>s+r.tn,0);
  html += `<tr style="background:#1a365d;color:#fff;font-weight:700;font-size:.92rem">
    <td style="position:sticky;left:0;z-index:2;background:#1a365d;color:#fff" colspan="2">Grand Total</td>
    <td style="position:sticky;left:190px;z-index:2;background:#1a365d;color:#fff">Net</td>
    <td style="text-align:right">${{grandNet}}</td>`;
  labels.forEach((_,i) => {{
    const v = rows.reduce((s,r)=>s+r.n[i],0);
    html += `<td style="text-align:right">${{v||''}}</td>`;
  }});
  html += '</tr>';
  tbody.innerHTML = html;
}}

function buildStartsTable(headId, bodyId, months, dates, rows) {{
  const thead = document.getElementById(headId);
  const tbody = document.getElementById(bodyId);
  const metricColors = {{
    'Completed Homes':'#003552','Homes Sold':'#276749','Cancellations':'#c53030',
    'Net Sales':'#276749','Under Construction':'#d69e2e','Vacant Lots':'#718096','Total Inventory':'#2d3748'
  }};
  const metricBg = {{'Total Inventory':'#edf2f7','Net Sales':'#f0fff4'}};
  const H = 'background:#003552;color:#fff';
  let th = `<tr>`
    + `<th style="position:sticky;left:0;z-index:3;${{H}};min-width:50px">Lot</th>`
    + `<th style="position:sticky;left:50px;z-index:3;${{H}};min-width:140px">Builder</th>`
    + `<th style="position:sticky;left:190px;z-index:3;${{H}};min-width:80px">Section</th>`
    + `<th style="position:sticky;left:270px;z-index:3;${{H}};min-width:80px">Delivery</th>`
    + `<th style="${{H}};min-width:130px">Metric</th>`;
  months.forEach((m,i) => th += `<th style="${{H}};min-width:80px;text-align:center">${{m}}<br><span style="font-weight:400;font-size:.75rem;opacity:.8">${{dates[i]}}</span></th>`);
  th += '</tr>';
  thead.innerHTML = th;
  let html = '';
  let prevLot = null, prevBld = null;
  rows.forEach(r => {{
    const bg = metricBg[r.metric] || '#fff';
    const clr = metricColors[r.metric] || '#2d3748';
    const isBldStart = r.lot !== prevLot || r.bld !== prevBld;
    html += `<tr style="background:${{bg}};border-bottom:1px solid #e2e8f0">`;
    html += `<td style="position:sticky;left:0;z-index:2;background:${{bg}};font-weight:700;text-align:center">${{isBldStart?r.lot+'ft':''}}</td>`;
    html += `<td style="position:sticky;left:50px;z-index:2;background:${{bg}};font-size:.8rem">${{isBldStart?r.bld:''}}</td>`;
    html += `<td style="position:sticky;left:190px;z-index:2;background:${{bg}};font-size:.78rem;color:#003552;font-style:italic">${{r.sec||''}}</td>`;
    html += `<td style="position:sticky;left:270px;z-index:2;background:${{bg}};font-size:.78rem;color:#003552;font-style:italic">${{r.del||''}}</td>`;
    html += `<td style="color:${{clr}};font-size:.8rem;font-weight:${{r.metric==='Total Inventory'||r.metric==='Net Sales'?'700':'400'}}">${{r.metric}}</td>`;
    r.vals.forEach(v => {{
      const numV = v === '' ? null : Number(v);
      const c = numV === null ? '' : (numV < 0 ? 'color:#c53030' : (numV > 0 && (r.metric==='Homes Sold'||r.metric==='Net Sales') ? 'color:#276749' : 'color:#2d3748'));
      html += `<td style="text-align:right;${{c}}">${{v !== '' ? v : ''}}</td>`;
    }});
    html += '</tr>';
    prevLot = r.lot; prevBld = r.bld;
  }});
  tbody.innerHTML = html;
}}

// ── RENDER EVERYTHING ─────────────────────────────────────────────────────────
buildWeeklyChart('wkGpdChart', gpdWkLabels, gpdWkGross, gpdWkCanc, gpdWkNet, '#003552', 'rgba(0,53,82,.18)');
buildWeeklyChart('wkHldChart', hldWkLabels, hldWkGross, hldWkCanc, hldWkNet, '#38a169', 'rgba(56,161,105,.18)');
buildWeeklyChart('wkWrgChart', wrgWkLabels, wrgWkGross, wrgWkCanc, wrgWkNet, '#e53e3e', 'rgba(229,62,62,.18)');

buildWeeklyDetailTable('wkGpdHead','wkGpdBody', gpdWkLabels, gpdWkRows);
buildWeeklyDetailTable('wkHldHead','wkHldBody', hldWkLabels, hldWkRows);
buildWeeklyDetailTable('wkWrgHead','wkWrgBody', wrgWkLabels, wrgWkRows);

buildStartsTable('wkGpdStartsHead','wkGpdStartsBody', gpdStartsMonths, gpdStartsDates, gpdStartsRows);
buildStartsTable('wkWrgStartsHead','wkWrgStartsBody', wrgStartsMonths, wrgStartsDates, wrgStartsRows);

// KPI totals
const gpdWkNetTotal = gpdWkNet.reduce((a,b)=>a+b,0);
const hldWkNetTotal = hldWkNet.reduce((a,b)=>a+b,0);
const wrgWkNetTotal = wrgWkNet.reduce((a,b)=>a+b,0);
document.getElementById('wk-gpd-total').textContent = gpdWkNetTotal;
document.getElementById('wk-hld-total').textContent = hldWkNetTotal;
document.getElementById('wk-wrg-total').textContent = wrgWkNetTotal;
document.getElementById('wk-combined-total').textContent = gpdWkNetTotal + hldWkNetTotal + wrgWkNetTotal;
</script>
</body>
</html>'''

    return html


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 3:
        print("Usage: python3 update_dashboard.py <excel_file> <data_date> [output_dir]")
        print('Example: python3 update_dashboard.py "data.xlsx" "03/30/2026" "./Sales Dashboard - Web"')
        sys.exit(1)

    excel_file = sys.argv[1]
    data_date = sys.argv[2]  # MM/DD/YYYY
    output_dir = sys.argv[3] if len(sys.argv) > 3 else '.'

    current_year = int(data_date.split('/')[-1])
    current_month = int(data_date.split('/')[0])

    print(f"Loading {excel_file}...")
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    print("Extracting Gross Sales TGP...")
    gpd_gross = extract_gross_sales(wb['Gross Sales TGP'])
    print(f"  {len(gpd_gross)} gross sales records")

    print("Extracting Cancellations TGP...")
    gpd_cancel = extract_cancellations(wb['Cancellations TGP'])
    print(f"  {len(gpd_cancel)} cancellation records")

    print("Extracting Gross Sales Highlands...")
    hld_gross = extract_gross_sales(wb['Gross Sales Highlands'])
    print(f"  {len(hld_gross)} gross sales records")

    print("Extracting Cancellations Highlands...")
    hld_cancel = extract_cancellations(wb['Cancellations Highlands'])
    print(f"  {len(hld_cancel)} cancellation records")

    print("Extracting Gross Sales WRG...")
    wrg_gross = extract_gross_sales(wb['Gross Sales WRG'])
    print(f"  {len(wrg_gross)} gross sales records")

    print("Extracting Cancellations WRG...")
    wrg_cancel = extract_cancellations(wb['Cancellations WRG'])
    print(f"  {len(wrg_cancel)} cancellation records")

    # Month ranges
    gpd_months = generate_month_range(TGP_START, (current_year, current_month))
    hld_months = generate_month_range(HLD_START, (current_year, current_month))
    wrg_months = generate_month_range((2024, 1), (current_year, current_month))

    # Monthly net sales
    gpd_net_monthly = compute_monthly_net(gpd_gross, gpd_cancel, gpd_months)
    hld_net_monthly = compute_monthly_net(hld_gross, hld_cancel, hld_months)
    wrg_net_monthly = compute_monthly_net(wrg_gross, wrg_cancel, wrg_months)

    # Builder-month net
    gpd_bm_net = compute_builder_month_net(gpd_gross, gpd_cancel, gpd_months)
    hld_bm_net = compute_builder_month_net(hld_gross, hld_cancel, hld_months)
    wrg_bm_net = compute_builder_month_net(wrg_gross, wrg_cancel, wrg_months)

    # All-time totals
    gpd_at = compute_all_time_totals(gpd_gross, gpd_cancel)
    hld_at = compute_all_time_totals(hld_gross, hld_cancel)
    wrg_at = compute_all_time_totals(wrg_gross, wrg_cancel)

    # Total net sales
    gpd_total_net = sum(info['tn'] for info in gpd_at.values())
    hld_total_net = sum(info['tn'] for info in hld_at.values())
    wrg_total_net = sum(info['tn'] for info in wrg_at.values())

    # YTD net
    ytd_months = generate_month_range((current_year, 1), (current_year, current_month))
    gpd_ytd_net = sum(gpd_net_monthly.get(m, 0) for m in ytd_months)
    hld_ytd_net = sum(hld_net_monthly.get(m, 0) for m in ytd_months)
    wrg_ytd_net = sum(wrg_net_monthly.get(m, 0) for m in ytd_months)

    # 2026 cancellation count
    gpd_canc_2026 = len([c for c in gpd_cancel if c['date'].year == current_year])
    hld_canc_2026 = len([c for c in hld_cancel if c['date'].year == current_year])
    wrg_canc_2026 = len([c for c in wrg_cancel if c['date'].year == current_year])

    # Weekly data
    gpd_weekly = compute_weekly_data(gpd_gross, gpd_cancel, current_year)
    hld_weekly = compute_weekly_data(hld_gross, hld_cancel, current_year)
    wrg_weekly = compute_weekly_data(wrg_gross, wrg_cancel, current_year)

    # Recent cancellations
    gpd_recent_cancels = sorted(gpd_cancel, key=lambda x: x['date'], reverse=True)[:5]

    # Totals for cancellation tab
    gpd_canc_total = len(gpd_cancel)
    hld_canc_total = len(hld_cancel)
    wrg_canc_total = len(wrg_cancel)
    gpd_gross_total = len(gpd_gross)
    hld_gross_total = len(hld_gross)
    wrg_gross_total = len(wrg_gross)

    # Read existing starts data from the current dashboard (if it exists)
    # Starts data is manually maintained and not in the gross/cancellations sheets
    gpd_starts_js = '[]'
    wrg_starts_js = '[]'
    existing_html_path = os.path.join(output_dir, 'net_sales_dashboard.html')
    if os.path.exists(existing_html_path):
        with open(existing_html_path, 'r') as f:
            existing_html = f.read()
        # Extract starts data
        import re
        m = re.search(r'const gpdStartsRows\s*=\s*(\[[\s\S]*?\]);', existing_html)
        if m:
            gpd_starts_js = m.group(1)
        m = re.search(r'const wrgStartsRows\s*=\s*(\[[\s\S]*?\]);', existing_html)
        if m:
            wrg_starts_js = m.group(1)

    print("\nData Summary:")
    print(f"  GPD: {gpd_total_net} net sales (LOP), {gpd_ytd_net} YTD {current_year}")
    print(f"  HLD: {hld_total_net} net sales (LOP), {hld_ytd_net} YTD {current_year}")
    print(f"  WRG: {wrg_total_net} net sales (LOP), {wrg_ytd_net} YTD {current_year}")
    print(f"  Combined GP: {gpd_total_net + hld_total_net}")

    data = {
        'gpd_months': gpd_months,
        'hld_months': hld_months,
        'wrg_months': wrg_months,
        'gpd_net_monthly': gpd_net_monthly,
        'hld_net_monthly': hld_net_monthly,
        'wrg_net_monthly': wrg_net_monthly,
        'gpd_bm_net': gpd_bm_net,
        'hld_bm_net': hld_bm_net,
        'wrg_bm_net': wrg_bm_net,
        'gpd_at': gpd_at,
        'hld_at': hld_at,
        'wrg_at': wrg_at,
        'gpd_total_net': gpd_total_net,
        'hld_total_net': hld_total_net,
        'wrg_total_net': wrg_total_net,
        'gpd_ytd_net': gpd_ytd_net,
        'hld_ytd_net': hld_ytd_net,
        'wrg_ytd_net': wrg_ytd_net,
        'gpd_canc_2026': gpd_canc_2026,
        'hld_canc_2026': hld_canc_2026,
        'wrg_canc_2026': wrg_canc_2026,
        'gpd_weekly': gpd_weekly,
        'hld_weekly': hld_weekly,
        'wrg_weekly': wrg_weekly,
        'gpd_recent_cancels': gpd_recent_cancels,
        'gpd_canc_total': gpd_canc_total,
        'hld_canc_total': hld_canc_total,
        'wrg_canc_total': wrg_canc_total,
        'gpd_gross_total': gpd_gross_total,
        'hld_gross_total': hld_gross_total,
        'wrg_gross_total': wrg_gross_total,
        'gpd_starts_js': gpd_starts_js,
        'wrg_starts_js': wrg_starts_js,
    }

    print("\nGenerating HTML...")
    html = generate_html(data, data_date)

    # Write output files
    out1 = os.path.join(output_dir, 'net_sales_dashboard.html')
    out2 = os.path.join(output_dir, 'index.html')

    with open(out1, 'w') as f:
        f.write(html)
    print(f"  Wrote {out1}")

    with open(out2, 'w') as f:
        f.write(html)
    print(f"  Wrote {out2}")

    print("\nDone! Dashboard updated successfully.")

if __name__ == '__main__':
    main()
